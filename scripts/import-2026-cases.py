#!/usr/bin/env python3
"""Build a transactional H2 import for the 2026 case registers.

The source workbooks are read-only. This script only creates review artifacts and
SQL; the caller is responsible for running the SQL against a stopped database.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BATCH = "LEGACY-2026-20260724"
RUN_AT = datetime(2026, 7, 24, 17, 0, 0)
DISABLED_PASSWORD = "$2a$10$Wq7piZaRUZNZfhbhh5Bf8esr2vxCuaxFfLPQTZjphUGY9GaEZR8le"

CASE_TYPE_BY_SHEET = {
    "民事": "CIVIL",
    "行政": "ADMINISTRATIVE",
    "非诉": "NON_LITIGATION",
    "顾问": "CONSULTANT",
    "刑事": "CRIMINAL",
}

CASE_CODE_BY_SHEET = {"民事": "民", "行政": "行", "非诉": "非", "顾问": "顾", "刑事": "刑"}

DEPARTMENT_ALIASES = {
    "金融": "金融事务部",
    "金融事务部": "金融事务部",
    "民商": "民商法务部",
    "民商法务部": "民商法务部",
    "清算重组": "重整与清算部",
    "清算重整": "重整与清算部",
    "破产清算部": "重整与清算部",
    "重整与清算部": "重整与清算部",
    "研究中心": "重大项目研究中心",
    "重大项目研究中心": "重大项目研究中心",
    "政府": "政府事务部",
    "政府事务部": "政府事务部",
    "法税": "法税合规部",
    "法税合规部": "法税合规部",
    "刑事": "刑事法律事务部",
    "刑事法律事务部": "刑事法律事务部",
    "房产": "房地产事务部",
    "房地产事务部": "房地产事务部",
    "主任": "主任室",
    "主任室": "主任室",
}

HISTORICAL_STAFF = {
    "吴泳琪": "金融事务部",
    "郑宇佳": "重整与清算部",
    "翁雪云": "政府事务部",
    "陈尚浩": "政府事务部",
}

NAME_FIXES = {
    "黄旑雯": "黄旖雯",
    "杨伟研": "杨伟妍",
    "陈  杰": "陈杰",
    "陈 杰": "陈杰",
}

AMBIGUOUS_KEYS = {"旧案？", "2026民007-015", "2026民006-022"}
SELECTED_DUPLICATES = {
    "2026民113": lambda row: row.sheet == "民事",
    "2026非010": lambda row: row.sheet == "非诉",
    "2026民057-008": lambda row: "何贤" in row.counterparty,
}


@dataclass
class RegisterRow:
    sheet: str
    excel_row: int
    filing_date: Any
    client_raw: str
    counterparty: str
    reason: str
    category: str
    department_raw: str
    lawyers_raw: str
    amount_raw: Any
    procedure_raw: str
    fee_raw: Any
    received_raw: Any
    contract_raw: str
    source_allocation_raw: str
    department_ratio_raw: Any
    firm_ratio_raw: Any
    notes_raw: str
    archive_status_raw: str
    archive_date_raw: Any
    service_start_raw: Any = None
    service_end_raw: Any = None
    business_type_raw: str = ""
    subject_matter_raw: str = ""

    @property
    def key(self) -> str:
        return normalize_case_key(self.contract_raw)


@dataclass
class ImportCase:
    register: RegisterRow
    case_list: dict[str, Any] | None
    case_number: str
    case_id: int = 0
    client_ids: list[int] = field(default_factory=list)
    party_rows: list[dict[str, Any]] = field(default_factory=list)
    owner_id: int = 0
    member_ids: list[int] = field(default_factory=list)
    department_id: int = 0
    department_name: str = ""
    lawyer_names: list[str] = field(default_factory=list)
    source_names: list[str] = field(default_factory=list)
    merged_existing: bool = False


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", "", text(value))


def normalize_name(value: Any) -> str:
    value = normalize_space(value)
    value = value.replace("（", "(").replace("）", ")")
    return value.strip("，,。.;；、")


def normalize_case_key(value: Any) -> str:
    raw = normalize_space(value)
    formal = re.fullmatch(r"\[(\d{4})\]粤至高([民行非顾刑])字第(\d{3})(?:-(\d+))?号", raw)
    if formal:
        suffix = f"-{int(formal.group(4)):03d}" if formal.group(4) else ""
        return f"{formal.group(1)}{formal.group(2)}{formal.group(3)}{suffix}"
    short = re.fullmatch(r"(\d{4})([民行非顾刑])(\d{3})(?:-(\d+))?", raw)
    if short:
        suffix = f"-{int(short.group(4)):03d}" if short.group(4) else ""
        return f"{short.group(1)}{short.group(2)}{short.group(3)}{suffix}"
    return raw


def formal_case_number(key: str) -> str:
    match = re.fullmatch(r"(\d{4})([民行非顾刑])(\d{3})(?:-(\d+))?", key)
    if not match:
        raise ValueError(f"Cannot format case number: {key}")
    suffix = f"-{int(match.group(4))}" if match.group(4) else ""
    return f"[{match.group(1)}]粤至高{match.group(2)}字第{match.group(3)}{suffix}号"


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    if not raw:
        return None
    for pattern in (r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})",):
        match = re.search(pattern, raw)
        if match:
            try:
                return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            except ValueError:
                return None
    return None


def decimal(value: Any) -> Decimal | None:
    if value is None or text(value) == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    match = re.search(r"-?\d+(?:\.\d+)?", text(value).replace(",", ""))
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def multiply_wan(value: Any) -> Decimal | None:
    amount = decimal(value)
    return amount * Decimal("10000") if amount is not None else None


def truncate(value: Any, limit: int) -> str | None:
    raw = text(value)
    return raw[:limit] if raw else None


def split_subjects(value: Any) -> list[str]:
    raw = text(value).replace("\r", "\n").replace("（", "(").replace("）", ")")
    if not raw:
        return []
    pieces = re.split(r"[、,，;；.。\n]+", raw)
    result = []
    for piece in pieces:
        cleaned = normalize_name(piece)
        cleaned = re.sub(r"\((?:原告|被告|受害人|被害人|委托人|申请人|被申请人|上诉人|被上诉人)\)$", "", cleaned)
        top_level = []
        main = []
        group = []
        depth = 0
        for char in cleaned:
            if char == "(":
                if depth == 0:
                    depth = 1
                    continue
                depth += 1
            elif char == ")" and depth:
                depth -= 1
                if depth == 0:
                    top_level.append("".join(group))
                    group = []
                    continue
            (group if depth else main).append(char)
        main_name = normalize_name("".join(main))
        alias = next((item for item in top_level if re.search(r"(?:改名为|更名为|原名|曾用名)[:：]?", item)), None)
        has_suffix_after_group = bool(top_level) and cleaned.rfind(")") < len(cleaned) - 1
        if alias:
            alias_name = normalize_name(re.sub(r"^.*?(?:改名为|更名为|原名|曾用名)[:：]?", "", alias))
            candidates = [alias_name or main_name]
        elif has_suffix_after_group or (top_level and all(item in {"有限合伙", "集团"} for item in top_level)):
            candidates = [cleaned]
        elif main_name and any(token in main_name for token in ("人民政府", "司法局", "公安局", "管理局", "委员会", "办公室")):
            candidates = [cleaned]
        elif top_level:
            candidates = [main_name] + [normalize_name(item) for item in top_level]
        else:
            candidates = [main_name]
        for candidate in candidates:
            if candidate and candidate not in result:
                result.append(candidate[:100])
    return result


def sql(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, Decimal)):
        return str(value)
    if isinstance(value, date) and not isinstance(value, datetime):
        return f"DATE '{value.isoformat()}'"
    if isinstance(value, datetime):
        return f"TIMESTAMP '{value.strftime('%Y-%m-%d %H:%M:%S')}'"
    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def read_case_list(path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    by_key: dict[str, dict[str, Any]] = {}
    all_rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(rows, 2):
        if not any(value not in (None, "") for value in values):
            continue
        row = dict(zip(headers, values))
        row["_excel_row"] = excel_row
        key = normalize_case_key(row.get("合同号"))
        if not key:
            continue
        row["_key"] = key
        if key in by_key:
            raise ValueError(f"案件清单编号重复: {key}")
        by_key[key] = row
        all_rows.append(row)
    return by_key, all_rows


def read_register(path: Path) -> list[RegisterRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: list[RegisterRow] = []
    for sheet in workbook.worksheets:
        if sheet.title not in CASE_TYPE_BY_SHEET:
            continue
        rows = sheet.iter_rows(values_only=True)
        next(rows)
        next(rows)
        for excel_row, values in enumerate(rows, 3):
            values = list(values)
            if len(values) < 2 or not text(values[1]):
                continue
            if sheet.title in {"民事", "行政"}:
                row = RegisterRow(sheet.title, excel_row, values[0], text(values[1]), text(values[2]), text(values[3]),
                                  text(values[4]), text(values[5]), text(values[6]), values[7], text(values[8]),
                                  values[17], values[18], text(values[23]), text(values[25] if sheet.title == "民事" else values[26]),
                                  values[26] if sheet.title == "民事" else values[27], values[27] if sheet.title == "民事" else values[28],
                                  text(values[28] if sheet.title == "民事" else values[32]), text(values[29]), values[30])
            elif sheet.title == "非诉":
                row = RegisterRow(sheet.title, excel_row, values[0], text(values[1]), "", text(values[3]), text(values[4]),
                                  text(values[5]), text(values[6]), values[7], "", values[12], values[13], text(values[17]),
                                  text(values[18]), values[19], values[20], text(values[23]), text(values[24]), values[25],
                                  business_type_raw=text(values[4]), subject_matter_raw=text(values[2]))
            elif sheet.title == "顾问":
                row = RegisterRow(sheet.title, excel_row, values[0], text(values[1]), "", "法律顾问", text(values[20]),
                                  text(values[4]), text(values[5]), None, "", values[10], values[11], text(values[15]),
                                  text(values[17]), values[18], values[19], text(values[22]), text(values[23]), values[24],
                                  service_start_raw=values[2], service_end_raw=values[3], business_type_raw="常年法律顾问")
            else:
                row = RegisterRow(sheet.title, excel_row, values[0], text(values[1]), text(values[2]), text(values[3]),
                                  "", text(values[4]), text(values[5]), None, text(values[6]), values[17], values[18],
                                  text(values[23]), text(values[26]), values[27], values[28], text(values[29]), text(values[30]), values[31],
                                  subject_matter_raw=text(values[2]))
            result.append(row)
    return result


def select_register_rows(rows: list[RegisterRow]) -> tuple[list[RegisterRow], list[dict[str, Any]]]:
    conflicts: list[dict[str, Any]] = []
    groups: dict[str, list[RegisterRow]] = defaultdict(list)
    for row in rows:
        if not row.key:
            conflicts.append(conflict(row, "缺少合同号，待行政补充"))
        else:
            groups[row.key].append(row)

    selected: list[RegisterRow] = []
    for key, group in groups.items():
        if key in AMBIGUOUS_KEYS:
            for row in group:
                conflicts.append(conflict(row, "编号为空或同号对应不同案件，暂缓导入"))
            continue
        if len(group) == 1:
            selected.append(group[0])
            continue
        chooser = SELECTED_DUPLICATES.get(key)
        if chooser:
            chosen = [row for row in group if chooser(row)]
            if len(chosen) == 1:
                selected.append(chosen[0])
                for row in group:
                    if row is not chosen[0]:
                        conflicts.append(conflict(row, f"与 {key} 重复，按确认方案合并或暂缓"))
                continue
        for row in group:
            conflicts.append(conflict(row, "登记表内部重复编号且无法自动选择，暂缓导入"))
    selected.sort(key=lambda row: (row.key, row.sheet, row.excel_row))
    return selected, conflicts


def conflict(row: RegisterRow, reason: str) -> dict[str, Any]:
    return {
        "来源": "2026年立案登记表.xlsx",
        "Sheet": row.sheet,
        "行号": row.excel_row,
        "合同号": row.contract_raw,
        "委托人": row.client_raw,
        "相对方/对象": row.counterparty or row.subject_matter_raw,
        "案由": row.reason,
        "主办律师": row.lawyers_raw,
        "暂缓原因": reason,
    }


def parse_staff(raw: Any, candidates: Iterable[str]) -> list[str]:
    value = text(raw)
    for old, new in NAME_FIXES.items():
        value = value.replace(old, new)
    compact = re.sub(r"[\s.。·、,，;；%％0-9]+", "", value)
    matches = []
    for name in set(candidates):
        if not name:
            continue
        for match in re.finditer(re.escape(name), compact):
            matches.append((match.start(), -len(name), match.end(), name))
    matches.sort()
    names = []
    occupied_until = -1
    for start, _, end, name in matches:
        if start < occupied_until or name in names:
            continue
        names.append(name)
        occupied_until = end
    return names


def department_name(raw: str, lawyer_names: list[str], users_by_name: dict[str, dict[str, Any]], departments_by_id: dict[int, str]) -> str:
    compact = normalize_space(raw)
    if compact == "二部" or not compact:
        for name in lawyer_names:
            user = users_by_name.get(name)
            if user and user.get("department_id"):
                return departments_by_id.get(user["department_id"], "")
    for alias in sorted(DEPARTMENT_ALIASES, key=len, reverse=True):
        if alias in compact and alias != "二部":
            return DEPARTMENT_ALIASES[alias]
    for name in lawyer_names:
        user = users_by_name.get(name)
        if user and user.get("department_id"):
            return departments_by_id.get(user["department_id"], "")
    return "广东至高律师事务所"


def client_type(name: str, source_value: Any = None) -> str:
    source = text(source_value)
    if source in {"个人", "企业", "金融机构", "事业单位", "党政机关", "社会团体", "其他"}:
        return source
    if any(token in name for token in ("银行", "金融资产", "证券", "保险")):
        return "金融机构"
    if any(token in name for token in ("人民政府", "司法局", "公安局", "法院", "检察院", "税务局", "委员会", "管理局", "街道办", "镇政府")):
        return "党政机关"
    if any(token in name for token in ("公司", "集团", "企业", "厂", "商行", "经营部", "合作社", "事务所", "中心", "医院", "学校", "协会", "村委会", "组织")):
        return "企业"
    return "个人"


def party_type(name: str) -> str:
    return "ORGANIZATION" if client_type(name) != "个人" else "INDIVIDUAL"


def party_roles(case_type: str, category: str) -> tuple[str, str]:
    value = text(category)
    if case_type == "CONSULTANT":
        return "CONSULTANT_UNIT", "SERVICE_RECIPIENT"
    if case_type == "NON_LITIGATION":
        return "CLIENT", "COUNTERPARTY"
    if case_type == "CRIMINAL":
        return "CLIENT", "SUSPECT"
    if "被告" in value or "被申请" in value:
        return "DEFENDANT", "PLAINTIFF"
    if "原告" in value:
        return "PLAINTIFF", "DEFENDANT"
    if "申请" in value:
        return "APPLICANT", "RESPONDENT"
    return "CLIENT", "COUNTERPARTY"


def fee_values(register_value: Any, case_list: dict[str, Any] | None) -> tuple[str, Decimal | None, Decimal | None, Decimal | None, str | None]:
    raw = text(register_value)
    list_method = text(case_list.get("收费方式")) if case_list else ""
    combined = raw or list_method
    risk_ratio = None
    ratio_match = re.search(r"(?:风险)?\s*(\d+(?:\.\d+)?)\s*[%％]", combined)
    if ratio_match:
        risk_ratio = Decimal(ratio_match.group(1))
    elif case_list:
        risk_ratio = decimal(case_list.get("风险比例（%）"))
    risk_fee = decimal(case_list.get("风险费用（元）")) if case_list else None

    numbers = [Decimal(value) for value in re.findall(r"\d+(?:\.\d+)?", combined.replace(",", ""))]
    if risk_ratio is not None and numbers and numbers[-1] == risk_ratio:
        numbers = numbers[:-1]
    fixed = None
    if isinstance(register_value, (int, float, Decimal)):
        fixed = decimal(register_value)
    elif numbers and ("风险" not in combined or "+" in combined or "加" in combined):
        fixed = sum(numbers, Decimal("0"))
    if fixed is None and case_list:
        fixed = decimal(case_list.get("固定费用（元）")) or decimal(case_list.get("合同金额（元）"))

    if "免费" in combined:
        method = "FREE"
    elif "风险" in combined and fixed:
        method = "FIXED_RISK"
    elif "风险" in combined:
        method = "RISK"
    elif fixed is not None or "固定" in combined:
        method = "FIXED"
    else:
        method = "OTHER"
    return method, fixed, risk_ratio, risk_fee, truncate(raw or list_method, 255)


def case_status(row: RegisterRow, list_row: dict[str, Any] | None) -> tuple[str, str | None]:
    if list_row and text(list_row.get("是否归档")) == "是":
        return "ARCHIVED", None
    if "归档" in row.archive_status_raw and "未" not in row.archive_status_raw:
        return "ARCHIVED", None
    if list_row and text(list_row.get("是否结案")) == "是":
        return "CLOSED", None
    if list_row and text(list_row.get("案件状态")) == "终止委托":
        return "CLOSED", "终止委托"
    return "ACTIVE", None


def build_case_name(row: RegisterRow, list_row: dict[str, Any] | None) -> str:
    if list_row and text(list_row.get("案件名称")):
        return truncate(list_row.get("案件名称"), 255) or ""
    if row.sheet == "顾问":
        return truncate(f"{row.client_raw}{row.business_type_raw or '法律顾问'}", 255) or ""
    target = row.counterparty or row.subject_matter_raw
    if target:
        return truncate(f"{row.client_raw} Vs {target} {row.reason}", 255) or ""
    return truncate(f"{row.client_raw} {row.reason}", 255) or ""


def choose_existing_client(name: str, department_id: int, existing: dict[str, list[dict[str, Any]]], decisions: list[dict[str, Any]]) -> int | None:
    candidates = existing.get(normalize_name(name), [])
    if not candidates:
        return None
    active = [item for item in candidates if not item["deleted"]] or candidates
    same_department = [item for item in active if item["department_id"] == department_id]
    chosen = min(same_department or active, key=lambda item: item["id"])
    if len(candidates) > 1:
        decisions.append({"客户名称": name, "候选客户ID": ",".join(str(item["id"]) for item in candidates),
                          "采用客户ID": chosen["id"], "规则": "优先同部门，其次采用最早有效记录"})
    return chosen["id"]


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title[:31])
        headers = list(rows[0].keys()) if rows else ["说明"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(45, max(12, max(len(text(cell.value)) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
            for cell in column:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-list", required=True, type=Path)
    parser.add_argument("--register", required=True, type=Path)
    parser.add_argument("--snapshot-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    case_list_by_key, case_list_rows = read_case_list(args.case_list)
    register_rows = read_register(args.register)
    selected_rows, conflicts = select_register_rows(register_rows)
    selected_keys = {row.key for row in selected_rows}
    for row in case_list_rows:
        if row["_key"] not in selected_keys:
            conflicts.append({"来源": "2026年案件清单.xlsx", "Sheet": "Sheet1", "行号": row["_excel_row"],
                              "合同号": row.get("合同号"), "委托人": row.get("委托人"),
                              "相对方/对象": row.get("对方当事人"), "案由": row.get("案由"),
                              "主办律师": row.get("主办律师"), "暂缓原因": "登记表无对应有效记录，按登记表为准暂缓导入"})

    department_rows = read_csv(args.snapshot_dir / "zgai-import-departments.csv")
    user_rows = read_csv(args.snapshot_dir / "zgai-import-users.csv")
    client_rows = read_csv(args.snapshot_dir / "zgai-import-clients.csv")
    case_rows = read_csv(args.snapshot_dir / "zgai-import-cases.csv")

    departments_by_name: dict[str, dict[str, Any]] = {}
    departments_by_id: dict[int, str] = {}
    for item in department_rows:
        department_id = int(item["ID"])
        name = item["DEPT_NAME"]
        record = {"id": department_id, "name": name, "leader_id": int(item["LEADER_ID"]) if item["LEADER_ID"] else None}
        departments_by_name[name] = record
        departments_by_id[department_id] = name

    users_by_name: dict[str, dict[str, Any]] = {}
    max_user_id = 0
    for item in user_rows:
        user_id = int(item["ID"])
        max_user_id = max(max_user_id, user_id)
        record = {"id": user_id, "name": item["REAL_NAME"], "username": item["USERNAME"],
                  "department_id": int(item["DEPARTMENT_ID"]) if item["DEPARTMENT_ID"] else None,
                  "status": int(item["STATUS"]), "deleted": item["DELETED"] == "TRUE"}
        existing = users_by_name.get(record["name"])
        if existing is None or (record["status"] == 1 and not record["deleted"]):
            users_by_name[record["name"]] = record

    historical_sql: list[str] = []
    for name, dept_name in HISTORICAL_STAFF.items():
        if name in users_by_name:
            continue
        max_user_id += 1
        dept = departments_by_name[dept_name]
        username = f"legacy_{hashlib.sha1(name.encode('utf-8')).hexdigest()[:12]}"
        record = {"id": max_user_id, "name": name, "username": username,
                  "department_id": dept["id"], "status": 0, "deleted": False}
        users_by_name[name] = record
        values = [max_user_id, RUN_AT, RUN_AT, False, username, DISABLED_PASSWORD, name, dept["id"], "历史人员", 0, False]
        historical_sql.append("INSERT INTO \"user\" (ID,CREATED_AT,UPDATED_AT,DELETED,USERNAME,PASSWORD,REAL_NAME,DEPARTMENT_ID,POSITION,STATUS,MUST_CHANGE_PASSWORD) VALUES (" + ",".join(sql(v) for v in values) + ");")

    existing_clients: dict[str, list[dict[str, Any]]] = defaultdict(list)
    max_client_id = 0
    for item in client_rows:
        client_id = int(item["ID"])
        max_client_id = max(max_client_id, client_id)
        existing_clients[normalize_name(item["CLIENT_NAME"])].append({
            "id": client_id, "department_id": int(item["DEPARTMENT_ID"]) if item["DEPARTMENT_ID"] else None,
            "deleted": item["DELETED"] == "TRUE",
        })

    max_case_id = max(int(item["ID"]) for item in case_rows)
    max_party_id = 15
    max_member_id = 11
    client_sql: list[str] = []
    case_sql: list[str] = []
    party_sql: list[str] = []
    member_sql: list[str] = []
    decisions: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []
    imported: list[ImportCase] = []
    staff_candidates = set(users_by_name)

    for register in selected_rows:
        list_row = case_list_by_key.get(register.key)
        case_number = text(list_row.get("合同号")) if list_row else formal_case_number(register.key)
        lawyer_names = parse_staff(register.lawyers_raw, staff_candidates)
        source_raw = register.source_allocation_raw
        if list_row and text(list_row.get("案源人")):
            source_raw = f"{source_raw} {list_row.get('案源人')}"
        source_names = parse_staff(source_raw, staff_candidates)
        dept_name = department_name(register.department_raw, lawyer_names, users_by_name, departments_by_id)
        dept = departments_by_name.get(dept_name) or departments_by_name["广东至高律师事务所"]
        active_lawyers = [users_by_name[name] for name in lawyer_names if name in users_by_name]
        owner = active_lawyers[0] if active_lawyers else None
        if owner is None:
            fallback_id = dept.get("leader_id") or 1
            owner = next((item for item in users_by_name.values() if item["id"] == fallback_id), users_by_name["系统管理员"])

        merged_existing = register.key == "2026顾071"
        if merged_existing:
            existing_case = next((item for item in case_rows if "洛伽建材" in item["CASE_NAME"]), None)
            if not existing_case:
                raise ValueError("未找到需要合并的洛伽建材顾问案件")
            case_id = int(existing_case["ID"])
        else:
            max_case_id += 1
            case_id = max_case_id

        import_case = ImportCase(register, list_row, case_number, case_id=case_id, owner_id=owner["id"],
                                 department_id=dept["id"], department_name=dept["name"], lawyer_names=lawyer_names,
                                 source_names=source_names, merged_existing=merged_existing)

        client_names = split_subjects(register.client_raw)
        if list_row:
            client_names.extend(name for name in split_subjects(list_row.get("其他委托人")) if name not in client_names)
        if not client_names:
            raise ValueError(f"案件 {register.key} 没有可导入的委托人")
        related_user_ids = [users_by_name[name]["id"] for name in source_names + lawyer_names if name in users_by_name]
        related_user_ids = list(dict.fromkeys(related_user_ids)) or [owner["id"]]
        for index, name in enumerate(client_names):
            existing_id = choose_existing_client(name, dept["id"], existing_clients, decisions)
            if existing_id is None:
                max_client_id += 1
                existing_id = max_client_id
                source_type = list_row.get("客户类型") if list_row and index == 0 else None
                values = [existing_id, RUN_AT, RUN_AT, False, name[:100], client_type(name, source_type), "ACTIVE",
                          f"{BATCH}:2026案件登记导入", f"来源案件 {case_number}", dept["id"], owner["id"],
                          ",".join(map(str, related_user_ids)), ",".join(map(str, related_user_ids))]
                client_sql.append("INSERT INTO CLIENT (ID,CREATED_AT,UPDATED_AT,DELETED,CLIENT_NAME,CLIENT_TYPE,STATUS,SOURCE,NOTES,DEPARTMENT_ID,OWNER_ID,SOURCE_USER_IDS,CLIENT_OWNER_IDS) VALUES (" + ",".join(sql(v) for v in values) + ");")
                existing_clients[normalize_name(name)].append({"id": existing_id, "department_id": dept["id"], "deleted": False})
            import_case.client_ids.append(existing_id)

        case_type = CASE_TYPE_BY_SHEET[register.sheet]
        client_role, counter_role = party_roles(case_type, register.category)
        for name in client_names:
            import_case.party_rows.append({"name": name, "role": client_role, "is_client": True})
        counter_names = split_subjects(register.counterparty or register.subject_matter_raw)
        for name in counter_names:
            if normalize_name(name) not in {normalize_name(value) for value in client_names}:
                import_case.party_rows.append({"name": name, "role": counter_role, "is_client": False})

        method, fixed_fee, risk_ratio, risk_fee, fee_notes = fee_values(register.fee_raw, list_row)
        amount = multiply_wan(register.amount_raw)
        if amount is None and list_row:
            amount = multiply_wan(list_row.get("标的（万元）（万元）"))
        filing_date = as_date(register.filing_date) or (as_date(list_row.get("立案日期（yyyy-MM-dd）")) if list_row else None)
        service_start = as_date(register.service_start_raw) or (as_date(list_row.get("合同服务开始时间（yyyy-MM-dd）")) if list_row else None)
        service_end = as_date(register.service_end_raw) or (as_date(list_row.get("合同服务结束时间（yyyy-MM-dd）")) if list_row else None)
        archive_date = as_date(register.archive_date_raw)
        status, close_status = case_status(register, list_row)
        summary_parts = []
        if list_row and text(list_row.get("案情介绍")):
            summary_parts.append(text(list_row.get("案情介绍")))
        if register.notes_raw:
            summary_parts.append(f"登记备注：{register.notes_raw}")
        summary = truncate("\n".join(summary_parts), 255)
        allocation = json.dumps({"batch": BATCH, "registerSheet": register.sheet, "registerRow": register.excel_row,
                                 "sourceAllocation": register.source_allocation_raw, "departmentRatio": text(register.department_ratio_raw),
                                 "firmRatio": text(register.firm_ratio_raw), "historicalLawyers": lawyer_names,
                                 "historicalSources": source_names}, ensure_ascii=False, separators=(",", ":"))
        generated_case_name = build_case_name(register, list_row)
        if merged_existing:
            generated_case_name = text(existing_case.get("CASE_NAME")) or generated_case_name
        case_values = {
            "ID": case_id, "CREATED_AT": RUN_AT, "UPDATED_AT": RUN_AT, "DELETED": False,
            "CASE_NUMBER": case_number, "CASE_NAME": generated_case_name, "CASE_TYPE": case_type,
            "CASE_REASON": truncate(register.reason or (list_row.get("案由") if list_row else ""), 100),
            "PROCEDURE": truncate(register.procedure_raw, 20), "LEVEL": "GENERAL", "STATUS": status,
            "CURRENT_STAGE": "历史案件导入", "COURT": truncate(list_row.get("审理法院") if list_row else None, 100),
            "ACCEPTANCE_DATE": filing_date, "FILING_DATE": filing_date, "COMMISSION_DATE": filing_date,
            "SUSPECT_NAME": truncate(register.counterparty if case_type == "CRIMINAL" else None, 100),
            "SUBJECT_MATTER": truncate(register.subject_matter_raw or (list_row.get("涉案主体/标的物") if list_row else None), 255),
            "BUSINESS_TYPE": truncate(register.business_type_raw or (list_row.get("业务类型") if list_row else None), 100),
            "AGENCY_TYPE": truncate(list_row.get("代理类型") if list_row else register.category, 50),
            "SERVICE_START_DATE": service_start, "SERVICE_END_DATE": service_end,
            "CONSULTANT_UNIT_NAME": truncate(register.client_raw if case_type == "CONSULTANT" else None, 200),
            "TRIAL_STAGES": truncate(register.procedure_raw, 255),
            "COURT_CASE_NUMBER": truncate(list_row.get("法院收案号") if list_row else None, 100),
            "HEARING_DATE": as_date(list_row.get("开庭时间（yyyy-MM-dd）")) if list_row else None,
            "CLOSE_DATE": archive_date if status in {"CLOSED", "ARCHIVED"} else None,
            "CLOSE_STATUS": close_status, "ARCHIVE_DATE": archive_date if status == "ARCHIVED" else None,
            "SUMMARY": summary, "TAGS": f"历史导入;2026台账;{register.sheet}", "AMOUNT": amount,
            "ATTORNEY_FEE": fixed_fee, "FEE_METHOD": method, "RISK_RATIO": risk_ratio, "RISK_FEE": risk_fee,
            "FEE_NOTES": fee_notes, "ALLOCATION_JSON": allocation,
            "ACTUAL_RECEIVED": decimal(register.received_raw), "OWNER_ID": owner["id"], "CLIENT_ID": import_case.client_ids[0],
        }
        columns = list(case_values)
        if merged_existing:
            assignments = [f"{column}={sql(value)}" for column, value in case_values.items() if column not in {"ID", "CREATED_AT", "DELETED"}]
            case_sql.append(f"UPDATE \"case\" SET {','.join(assignments)} WHERE ID={case_id};")
        else:
            case_sql.append("INSERT INTO \"case\" (" + ",".join(columns) + ") VALUES (" + ",".join(sql(case_values[column]) for column in columns) + ");")

        existing_merged_party_names = {normalize_name(register.client_raw)} if merged_existing else set()
        if merged_existing:
            party_sql.append("UPDATE PARTY SET PARTY_TYPE='ORGANIZATION',PARTY_ROLE='CONSULTANT_UNIT',IS_CLIENT=TRUE,UPDATED_AT=" +
                             sql(RUN_AT) + ",NOTES=" + sql(f"{BATCH} 合并保留") + f" WHERE CASE_ID={case_id} AND DELETED=FALSE;")
        for party in import_case.party_rows:
            if merged_existing and normalize_name(party["name"]) in existing_merged_party_names:
                continue
            max_party_id += 1
            values = [max_party_id, RUN_AT, RUN_AT, False, case_id, party_type(party["name"]), party["role"],
                      party["name"][:100], party["is_client"], f"{BATCH} 自动拆分"]
            party_sql.append("INSERT INTO PARTY (ID,CREATED_AT,UPDATED_AT,DELETED,CASE_ID,PARTY_TYPE,PARTY_ROLE,NAME,IS_CLIENT,NOTES) VALUES (" + ",".join(sql(v) for v in values) + ");")

        member_users = [users_by_name[name] for name in lawyer_names if name in users_by_name]
        if dept.get("leader_id") and dept["leader_id"] not in {user["id"] for user in member_users}:
            leader = next((user for user in users_by_name.values() if user["id"] == dept["leader_id"]), None)
            if leader:
                member_users.append(leader)
        existing_member_ids = {int(existing_case["OWNER_ID"])} if merged_existing and existing_case.get("OWNER_ID") else set()
        for index, user in enumerate(member_users):
            if user["id"] in existing_member_ids:
                continue
            existing_member_ids.add(user["id"])
            max_member_id += 1
            member_type = "CO_OWNER" if index else "OWNER"
            member_sql.append("INSERT INTO CASE_MEMBER (ID,CASE_ID,CREATED_AT,DELETED,MEMBER_TYPE,UPDATED_AT,USER_ID) VALUES (" +
                              ",".join(sql(v) for v in [max_member_id, case_id, RUN_AT, False, member_type, RUN_AT, user["id"]]) + ");")

        preview.append({"案件ID": case_id, "合同号": register.key, "正式编号": case_number,
                        "案件名称": case_values["CASE_NAME"], "类型": register.sheet, "部门": dept["name"],
                        "主办律师": "、".join(lawyer_names) or owner["name"], "委托人": "、".join(client_names),
                        "相对方/对象": register.counterparty or register.subject_matter_raw,
                        "收费原文": text(register.fee_raw), "状态": status, "处理方式": "合并现有案件" if merged_existing else "新建"})
        imported.append(import_case)

    statements = ["SET AUTOCOMMIT FALSE;",
                  "SELECT 1 / CASE WHEN (SELECT COUNT(*) FROM \"case\" WHERE DELETED=FALSE)=8 THEN 1 ELSE 0 END AS BASELINE_OK;"]
    statements.extend([
        "UPDATE \"case\" SET CASE_NUMBER=NULL,UPDATED_AT=" + sql(RUN_AT) + " WHERE ID IN (5,6,8);",
        *historical_sql, *client_sql, *case_sql, *party_sql, *member_sql,
        "INSERT INTO AUDIT_LOG (ID,USER_ID,MODULE,OPERATION,METHOD,PARAMS,IP,STATUS,EXECUTION_TIME,CREATED_AT) VALUES (50,1,'LEGACY_CASE_IMPORT','IMPORT_2026_CASES','OFFLINE_SQL'," +
        sql(json.dumps({"batch": BATCH, "imported": len(imported), "deferred": len(conflicts),
                        "caseListSha256": sha256(args.case_list), "registerSha256": sha256(args.register)}, ensure_ascii=False)) + ",'127.0.0.1',1,0," + sql(RUN_AT) + ");",
        "COMMIT;",
    ])
    (args.output_dir / "import.sql").write_text("\n".join(statements) + "\n", encoding="utf-8")
    identity_sql = [
        f'ALTER TABLE "user" ALTER COLUMN ID RESTART WITH {max_user_id + 1};',
        f"ALTER TABLE CLIENT ALTER COLUMN ID RESTART WITH {max_client_id + 1};",
        f'ALTER TABLE "case" ALTER COLUMN ID RESTART WITH {max_case_id + 1};',
        f"ALTER TABLE PARTY ALTER COLUMN ID RESTART WITH {max_party_id + 1};",
        f"ALTER TABLE CASE_MEMBER ALTER COLUMN ID RESTART WITH {max_member_id + 1};",
        "ALTER TABLE AUDIT_LOG ALTER COLUMN ID RESTART WITH 51;",
    ]
    (args.output_dir / "post-import-identity.sql").write_text("\n".join(identity_sql) + "\n", encoding="utf-8")
    write_workbook(args.output_dir / "2026案件导入预览与待补清单.xlsx", {
        "导入预览": preview,
        "待补与冲突": conflicts,
        "客户去重决策": decisions,
    })
    summary = {
        "batch": BATCH,
        "registerRows": len(register_rows),
        "selectedRegisterRows": len(selected_rows),
        "importedCases": len(imported),
        "newCases": sum(not item.merged_existing for item in imported),
        "mergedCases": sum(item.merged_existing for item in imported),
        "deferredRows": len(conflicts),
        "newClients": len(client_sql),
        "newParties": max_party_id - 15,
        "newCaseMembers": len(member_sql),
        "historicalStaff": len(historical_sql),
        "maxIds": {"user": max_user_id, "client": max_client_id, "case": max_case_id,
                   "party": max_party_id, "caseMember": max_member_id},
    }
    (args.output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
