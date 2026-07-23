#!/usr/bin/env python3
"""Generate the administrative workbook used to stage legacy case imports."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "ZGAI旧案资料导入模板_v1.xlsx"

BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_RED = "FCE4D6"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E1F2")


DICTIONARIES = {
    "案件类型": ["民事", "刑事", "行政", "非诉", "顾问"],
    "案件状态": ["咨询中", "待审批", "办理中", "已结案", "已归档"],
    "收费方式": ["固定收费", "风险收费", "基础+风险", "其他", "免费代理", "未确定"],
    "主体类型": ["个人", "企业", "其他组织"],
    "是否": ["是", "否"],
    "客户关系": ["委托人", "当事人", "联系人", "顾问单位", "其他"],
    "当事人角色": [
        "原告", "被告", "第三人", "申请人", "被申请人", "上诉人", "被上诉人",
        "再审申请人", "被申请执行人", "犯罪嫌疑人", "被告人", "被害人", "其他",
    ],
    "文档类型": ["立案材料", "证据材料", "法律文书", "合同收费", "往来函件", "审批归档", "其他"],
    "所属部门": [
        "金融事务部", "民商法务部", "重整与清算部", "重大项目研究中心", "政府事务部",
        "法税合规部", "刑事法律事务部", "独立律师团队", "后勤保障部",
    ],
}


CASE_HEADERS = [
    ("导入批次号", True, "同一次导入使用相同批次号，如 LEGACY-2026-01"),
    ("旧系统案件编号", True, "旧系统中的唯一编号；与导入批次号共同作为幂等键"),
    ("原ZGAI案件编号", False, "如历史资料已有本所编号则填写；没有可留空"),
    ("案件名称", True, "建议格式：客户/我方主体 Vs 相对方 + 案由"),
    ("案件类型", True, "民事、刑事、行政、非诉、顾问"),
    ("案由", False, "自由填写，不使用编号代替"),
    ("业务类型", False, "如金融、公司、婚姻家庭、刑事辩护等"),
    ("审级/程序", False, "多个值用中文分号分隔，如 一审；二审"),
    ("案件状态", True, "咨询中、待审批、办理中、已结案、已归档"),
    ("当前阶段", False, "按旧案实际办理阶段填写"),
    ("所属部门", True, "必须与系统部门名称完全一致"),
    ("案源人姓名", False, "多人用中文分号分隔，姓名须与员工账号一致"),
    ("主办律师姓名", True, "姓名须与员工账号一致"),
    ("协办人姓名", False, "多人用中文分号分隔"),
    ("客户名称", True, "须能在客户映射表找到；多个客户用中文分号分隔"),
    ("委托日期", False, "YYYY-MM-DD"),
    ("收案日期", False, "YYYY-MM-DD"),
    ("审理机构/法院", False, "法院、仲裁委、行政机关或其他受理机构"),
    ("法院/仲裁案号", False, "保持原格式"),
    ("立案日期", False, "YYYY-MM-DD"),
    ("开庭日期", False, "YYYY-MM-DD；多次开庭请在备注补充"),
    ("期限日期", False, "当前最重要期限，YYYY-MM-DD"),
    ("结案日期", False, "YYYY-MM-DD"),
    ("结案状态", False, "如判决、调解、撤诉、裁决、终止委托"),
    ("归档日期", False, "YYYY-MM-DD"),
    ("原归档位置", False, "纸质档案柜或旧电子目录位置"),
    ("涉案标的额(元)", False, "纯数字，不加逗号和货币符号"),
    ("律师费(元)", False, "纯数字"),
    ("收费方式", False, "固定收费、风险收费、基础+风险、其他、免费代理、未确定"),
    ("风险比例(%)", False, "填写 0-18 的数字"),
    ("风险收费金额(元)", False, "纯数字"),
    ("胜诉/实现债权金额(元)", False, "纯数字"),
    ("实际收款金额(元)", False, "纯数字"),
    ("代理类型", False, "如原告代理、被告代理、辩护人、专项顾问"),
    ("服务开始日期", False, "顾问/非诉项目使用，YYYY-MM-DD"),
    ("服务结束日期", False, "顾问/非诉项目使用，YYYY-MM-DD"),
    ("案情摘要", False, "建议 500 字以内，仅写客观事实与当前进展"),
    ("标签", False, "多个标签用中文分号分隔"),
    ("备注", False, "无法结构化的历史信息"),
]

PARTY_HEADERS = [
    ("导入批次号", True, "必须与案件主表一致"),
    ("旧系统案件编号", True, "必须与案件主表一致"),
    ("当事人序号", True, "同一案件从 1 开始递增"),
    ("当事人名称", True, "个人姓名或单位全称"),
    ("主体类型", True, "个人、企业、其他组织"),
    ("当事人角色", True, "原告、被告、申请人、被申请人等"),
    ("是否委托客户", True, "是/否"),
    ("性别", False, "个人主体填写"),
    ("民族", False, "个人主体填写"),
    ("身份证号", False, "按文本填写，避免科学计数法"),
    ("统一社会信用代码", False, "单位主体填写"),
    ("联系电话", False, "手机号或区号+固话"),
    ("地址", False, "联系地址/注册地址"),
    ("法定代表人", False, "单位主体填写"),
    ("对方律师", False, "已知时填写"),
    ("备注", False, "关联企业、曾用名等"),
]

CLIENT_HEADERS = [
    ("导入批次号", True, "必须与案件主表一致"),
    ("旧系统案件编号", True, "必须与案件主表一致"),
    ("客户名称", True, "个人姓名或单位全称"),
    ("客户类型", True, "个人、企业、其他组织"),
    ("客户关系", True, "委托人、当事人、联系人、顾问单位、其他"),
    ("身份证号", False, "个人主体填写"),
    ("统一社会信用代码", False, "单位主体填写；优先用于去重"),
    ("联系电话", False, "手机号或区号+固话"),
    ("联系地址", False, "可留空"),
    ("所属部门", True, "必须与案件主表一致或说明例外"),
    ("案源人姓名", False, "多人用中文分号分隔"),
    ("承办人姓名", True, "多人用中文分号分隔"),
    ("已存在于ZGAI", True, "是/否；不确定先填否，由导入预检判断"),
    ("备注", False, "曾用名、关联主体等"),
]

FILE_HEADERS = [
    ("导入批次号", True, "必须与案件主表一致"),
    ("旧系统案件编号", True, "必须与案件主表一致"),
    ("文件相对路径", True, "相对于旧案资料根目录；禁止填写磁盘绝对路径"),
    ("原文件名", True, "包含扩展名"),
    ("文档类型", False, "立案材料、证据材料、法律文书等"),
    ("建议案件目录", False, "确认后才写入一案一档"),
    ("文件日期", False, "YYYY-MM-DD"),
    ("是否允许进入知识库", True, "旧案件材料默认填否"),
    ("备注", False, "扫描件、重复件、缺页等说明"),
]


def add_sheet(workbook, title, headers):
    ws = workbook.create_sheet(title)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for col, (name, required, note) in enumerate(headers, 1):
        cell = ws.cell(1, col, f"{name}{'*' if required else ''}")
        cell.font = Font(color=WHITE, bold=True)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.comment = __import__("openpyxl").comments.Comment(note, "ZGAI")
        width = max(13, min(32, len(name) * 2 + 4))
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2, max_row=1001, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    required_cols = [i for i, (_, required, _) in enumerate(headers, 1) if required]
    for col in required_cols:
        letter = ws.cell(1, col).column_letter
        ws.conditional_formatting.add(
            f"{letter}2:{letter}1001",
            FormulaRule(formula=[f'AND(COUNTA($A2:${ws.cell(1, len(headers)).column_letter}2)>0,{letter}2="")'],
                        fill=PatternFill("solid", fgColor=LIGHT_RED)),
        )
    return ws


def add_list_validation(ws, header_name, formula):
    headers = {cell.value.rstrip("*"): cell.column for cell in ws[1]}
    column = headers.get(header_name)
    if not column:
        return
    letter = ws.cell(1, column).column_letter
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "请从下拉列表中选择标准值"
    validation.errorTitle = "值不符合导入字典"
    validation.prompt = "请使用系统标准值"
    validation.promptTitle = header_name
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}1001")


def build_workbook():
    wb = Workbook()
    guide = wb.active
    guide.title = "填写说明"
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 100
    guide.append(["项目", "填写要求"])
    instructions = [
        ("使用目的", "供行政人员整理旧系统案件、客户、当事人和电子文件索引；正式导入前由系统执行预检，不会直接覆盖现有数据。"),
        ("主键规则", "同一导入批次内，“旧系统案件编号”必须唯一。其他工作表用“导入批次号 + 旧系统案件编号”关联案件。"),
        ("必填规则", "表头带 * 的列为必填。只要某行开始填写，该行全部必填项必须补齐；缺失项会标红。"),
        ("姓名规则", "案源人、主办律师、协办人和承办人必须使用系统员工账号对应的真实姓名；多人用中文分号“；”分隔。"),
        ("日期与金额", "日期统一 YYYY-MM-DD。金额以元为单位，只填数字，不加逗号、货币符号或“万元”。"),
        ("隐私信息", "身份证号、统一社会信用代码、电话等列已按文本处理。工作簿只能在律所受控存储中传递。"),
        ("文件规则", "文件索引只填相对于旧案资料根目录的路径，不移动、不删除原文件。案件材料默认禁止进入共享知识库。"),
        ("导入顺序", "先预检员工/部门和客户去重，再导入案件主表与当事人，最后关联文件。任何错误均应整批回滚。"),
        ("禁止事项", "不得修改表头、合并数据单元格、使用同一旧案编号代表多个案件，或在同一单元格放多名当事人的完整资料。"),
    ]
    for item in instructions:
        guide.append(item)
    for cell in guide[1]:
        cell.font = Font(color=WHITE, bold=True)
        cell.fill = PatternFill("solid", fgColor=BLUE)
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    case_ws = add_sheet(wb, "案件主表", CASE_HEADERS)
    party_ws = add_sheet(wb, "当事人表", PARTY_HEADERS)
    client_ws = add_sheet(wb, "客户映射表", CLIENT_HEADERS)
    file_ws = add_sheet(wb, "文件索引表", FILE_HEADERS)

    dictionary = wb.create_sheet("字典")
    for col, (name, values) in enumerate(DICTIONARIES.items(), 1):
        dictionary.cell(1, col, name)
        dictionary.cell(1, col).font = Font(color=WHITE, bold=True)
        dictionary.cell(1, col).fill = PatternFill("solid", fgColor=BLUE)
        dictionary.column_dimensions[dictionary.cell(1, col).column_letter].width = 24
        for row, value in enumerate(values, 2):
            dictionary.cell(row, col, value)

    ranges = {}
    for col, (name, values) in enumerate(DICTIONARIES.items(), 1):
        letter = dictionary.cell(1, col).column_letter
        ranges[name] = f"'字典'!${letter}$2:${letter}${len(values) + 1}"

    for ws in (case_ws,):
        add_list_validation(ws, "案件类型", ranges["案件类型"])
        add_list_validation(ws, "案件状态", ranges["案件状态"])
        add_list_validation(ws, "收费方式", ranges["收费方式"])
        add_list_validation(ws, "所属部门", ranges["所属部门"])
    add_list_validation(party_ws, "主体类型", ranges["主体类型"])
    add_list_validation(party_ws, "当事人角色", ranges["当事人角色"])
    add_list_validation(party_ws, "是否委托客户", ranges["是否"])
    add_list_validation(client_ws, "客户类型", ranges["主体类型"])
    add_list_validation(client_ws, "客户关系", ranges["客户关系"])
    add_list_validation(client_ws, "所属部门", ranges["所属部门"])
    add_list_validation(client_ws, "已存在于ZGAI", ranges["是否"])
    add_list_validation(file_ws, "文档类型", ranges["文档类型"])
    add_list_validation(file_ws, "是否允许进入知识库", ranges["是否"])

    wb.save(OUTPUT)


def verify_workbook():
    wb = load_workbook(OUTPUT, data_only=False)
    expected = {"填写说明", "案件主表", "当事人表", "客户映射表", "文件索引表", "字典"}
    assert set(wb.sheetnames) == expected
    assert wb["案件主表"].max_column == len(CASE_HEADERS)
    assert wb["当事人表"].max_column == len(PARTY_HEADERS)
    assert wb["客户映射表"].max_column == len(CLIENT_HEADERS)
    assert wb["文件索引表"].auto_filter.ref == "A1:I1"


if __name__ == "__main__":
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    build_workbook()
    verify_workbook()
    print(OUTPUT)
