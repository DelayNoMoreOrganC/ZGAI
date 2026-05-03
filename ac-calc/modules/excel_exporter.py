"""
Excel 导出模块 v0.4
输出优化：4个Sheet，适配法院呈交格式。
"""
import io
from typing import Any, Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from modules.bank_loan_flow import TEMPLATE_KEY as FLOW_TEMPLATE_KEY


class ExcelExporter:
    """Excel 导出器 - 输出4个Sheet的完整债权计算表。"""

    def __init__(self):
        self.date_fmt = "YYYY-MM-DD"
        self.money_fmt = '#,##0.00'
        self.rate_fmt = '0.0000'
        self.header_font = Font(bold=True, size=10)
        self.title_font = Font(bold=True, size=14)
        self.total_font = Font(bold=True, size=10, color="FF0000")
        self.header_fill = PatternFill("solid", fgColor="D9E1F2")  # 浅蓝
        self.light_fill = PatternFill("solid", fgColor="F2F2F2")  # 浅灰
        self.thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        self.cell_align = Alignment(horizontal="right", vertical="center")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _apply_header(self, ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            c.font = self.header_font
            c.fill = self.header_fill
            c.alignment = self.center_align
            c.border = self.thin_border

    def _apply_money(self, ws, row, col, value):
        c = ws.cell(row, col, value)
        if isinstance(value, (int, float)) and value != 0:
            c.number_format = self.money_fmt
        c.border = self.thin_border
        c.alignment = self.cell_align
        return c

    def _apply_rate(self, ws, row, col, value):
        c = ws.cell(row, col, value)
        if isinstance(value, (int, float)) and value != 0:
            c.number_format = self.rate_fmt
        c.border = self.thin_border
        c.alignment = self.center_align
        return c

    # =================================================================
    # 导出入口
    # =================================================================
    def export_calculation_result(self, calc_result, loan_data, repayment_records=None):
        if calc_result.get("template_key") == FLOW_TEMPLATE_KEY:
            return self._export_loan_flow_template_result(calc_result, loan_data, repayment_records or [])

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Sheet 1: 债权汇总（最简概览）
        self._create_summary_sheet(wb, calc_result["summary"], loan_data, repayment_records or [])
        # Sheet 2: 本息计算表（法院呈交格式）
        self._create_detail_sheet(wb, calc_result, loan_data, repayment_records or [])
        # Sheet 3: 还款记录
        if repayment_records:
            self._create_records_sheet(wb, repayment_records, calc_result.get("repayment_details", []))
        # Sheet 4: 公式核对
        self._create_formula_sheet(wb, calc_result["detail"])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    # =================================================================
    # Sheet 1: 债权汇总
    # =================================================================
    def _create_summary_sheet(self, wb, summary, loan_data, repay_records):
        ws = wb.create_sheet("债权汇总")
        ws.sheet_properties.tabColor = "4472C4"

        # -------- 标题 --------
        calc_date = summary.get("calculation_date", "")
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "债权计算汇总表"
        c.font = Font(bold=True, size=16)
        c.alignment = self.center_align

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"数据暂计至：{calc_date}"
        ws["A2"].font = Font(size=10, italic=True)
        ws["A2"].alignment = self.center_align

        # -------- 债权概览 --------
        row = 4
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = "一、债权概览"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        row = 5
        for col, h in enumerate(["项目", "金额（元）", "说明"], 1):
            c = ws.cell(row, col, h)
            c.font = self.header_font; c.fill = self.header_fill
            c.alignment = self.center_align; c.border = self.thin_border
        for col in range(4, 7):
            ws.cell(row, col).fill = self.header_fill

        items = [
            ("贷款本金", loan_data.get("loan_amount", 0), "合同贷款金额"),
            ("剩余本金", summary.get("remaining_principal", 0), "扣除已还本金"),
            ("欠付利息", summary.get("accrued_interest", 0), "暂计日止应付未付利息"),
            ("欠付罚息", summary.get("penalty_interest", 0), "逾期本金按罚息利率计算"),
            ("欠付复利", summary.get("compound_interest", 0), "欠付利息按期计收复利"),
        ]

        for i, (label, val, note) in enumerate(items):
            r = row + 1 + i
            bg = self.light_fill if i % 2 == 0 else None
            ws.cell(r, 1, label).border = self.thin_border; ws.cell(r, 1).alignment = self.center_align
            if bg: ws.cell(r, 1).fill = bg
            c = self._apply_money(ws, r, 2, val)
            if bg: c.fill = bg
            ws.cell(r, 3, note).border = self.thin_border; ws.cell(r, 3).alignment = self.left_align
            if bg: ws.cell(r, 3).fill = bg

        # Total row
        total_row = row + 1 + len(items)
        ws.cell(total_row, 1, "债权总额").font = self.total_font
        ws.cell(total_row, 1).border = self.thin_border
        ws.cell(total_row, 1).alignment = self.center_align
        total_val = summary.get("total_amount", 0)
        c = ws.cell(total_row, 2, total_val)
        c.font = self.total_font
        c.number_format = self.money_fmt
        c.border = self.thin_border
        ws.cell(total_row, 3, "本金 + 利息 + 罚息 + 复利").border = self.thin_border

        # -------- 案件信息 --------
        row = total_row + 2
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = "二、案件信息"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        info_row = row + 1
        info_items = [
            ("贷款本金", f"{loan_data.get('loan_amount', 0):,.2f} 元"),
            ("放款日", loan_data.get("start_date", "")),
            ("合同到期日", loan_data.get("end_date", "")),
            ("数据暂计日", calc_date),
            ("计息基准", f"实际天数/{loan_data.get('day_base', 360)}"),
            ("罚息上浮", f"{loan_data.get('penalty_mult', 50)}%"),
        ]

        for i, (label, val) in enumerate(info_items):
            r = info_row + i
            ws.cell(r, 1, label).font = self.header_font
            ws.cell(r, 1).border = self.thin_border
            ws.cell(r, 1).alignment = self.center_align
            ws.cell(r, 2, val).border = self.thin_border
            ws.cell(r, 2).alignment = self.left_align
            ws.merge_cells(f"B{r}:F{r}")

        # -------- 还款流水统计 --------
        row = info_row + len(info_items) + 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = "三、还款流水"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        if repay_records:
            total_principal = sum(r["amount"] for r in repay_records if r.get("type", "") in ("principal", "本金"))
            total_interest = sum(r["amount"] for r in repay_records if r.get("type", "") in ("interest", "利息"))
            stats_row = row + 1
            ws.cell(stats_row, 1, "已还本金合计").border = self.thin_border
            ws.cell(stats_row, 1).alignment = self.center_align
            ws.cell(stats_row, 2, total_principal).border = self.thin_border
            ws.cell(stats_row, 2).number_format = self.money_fmt
            ws.cell(stats_row, 3, f"共 {len(repay_records)} 笔").border = self.thin_border
            ws.cell(stats_row, 3).alignment = self.left_align
            ws.merge_cells(f"C{stats_row}:F{stats_row}")
            stats_row += 1
            ws.cell(stats_row, 1, "已还利息合计").border = self.thin_border
            ws.cell(stats_row, 1).alignment = self.center_align
            ws.cell(stats_row, 2, total_interest).border = self.thin_border
            ws.cell(stats_row, 2).number_format = self.money_fmt
            ws.cell(stats_row, 3, "").border = self.thin_border

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

    # =================================================================
    # Sheet 2: 本息计算表（法院呈交格式）
    # =================================================================
    def _create_detail_sheet(self, wb, calc_result, loan_data, repay_records):
        ws = wb.create_sheet("本息计算表")
        ws.sheet_properties.tabColor = "E06C00"
        detail = calc_result.get("detail", [])
        summary = calc_result.get("summary", {})
        calc_date = summary.get("calculation_date", "")

        # -------- 表格标题 --------
        row = 1
        ws.merge_cells("A1:S1")
        ws["A1"].value = f"本息计算表（暂计至{calc_date}）"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_align

        # -------- 案件信息 --------
        row = 2
        ws.merge_cells("A2:S2")
        loan_amt = loan_data.get("loan_amount", 0)
        start = loan_data.get("start_date", "")
        end = loan_data.get("end_date", "")
        ws["A2"].value = f"贷款金额：{loan_amt:,.2f}元    借款期限：{start} 至 {end}"
        ws["A2"].font = Font(size=9)
        ws["A2"].alignment = self.left_align

        # -------- 表头 --------
        header_row = 4
        headers = [
            "序号", "本金余额", "起算日", "暂计至", "天数",
            "应还本金", "已还本金", "累计逾期本金",
            "正常贷款\n年利率", "应付利息", "已还利息", "累计尚欠利息",
            "罚息\n年利率", "应付罚息", "已还罚息", "累计尚欠\n罚息",
            "复利\n年利率", "应付复利", "累计尚欠\n复利",
        ]
        self._apply_header(ws, headers, header_row)

        # -------- 数据行 --------
        for i, item in enumerate(detail):
            r = header_row + 1 + i
            bg = self.light_fill if i % 2 == 1 else None

            row_data = [
                (item.get("seq_no"), "center", None),
                (item.get("principal_balance", 0), "money", None),
                (item.get("start_date", ""), "center", None),
                (item.get("end_date", ""), "center", None),
                (item.get("days", 0), "center", None),
                (item.get("principal_due", 0), "money", None),
                (item.get("principal_repaid", 0), "money", None),
                (item.get("cumulative_overdue_principal", 0), "money", None),
                (item.get("normal_rate", None), "rate", None),
                (item.get("new_interest", 0), "money", None),
                (item.get("repaid_interest", 0), "money", None),
                (item.get("interest_balance", 0), "money", None),
                (item.get("penalty_rate", None), "rate", None),
                (item.get("new_penalty", 0), "money", None),
                (item.get("repaid_penalty", 0), "money", None),
                (item.get("penalty_balance", 0), "money", None),
                (item.get("compound_rate", None), "rate", None),
                (item.get("new_interest_compound", 0), "money", None),
                (item.get("interest_compound_balance", 0), "money", None),
            ]
            # Build rates from detail
            # 正常贷款年利率: use the first adjustment rate or 0
            normal_rate = item.get("normal_rate", 0)
            if not normal_rate or normal_rate <= 0:
                normal_rate = 0
            penalty_rate = normal_rate * 1.5 if normal_rate else 0
            if item.get("penalty_balance", 0) > 0:
                compound_rate = penalty_rate
            else:
                compound_rate = normal_rate

            row_data[8] = (normal_rate, "rate", None)
            row_data[12] = (penalty_rate, "rate", None)
            row_data[16] = (compound_rate, "rate", None)

            for col, (val, fmt, _) in enumerate(row_data, 1):
                c = ws.cell(r, col)
                c.border = self.thin_border
                if bg:
                    c.fill = bg
                if fmt == "money":
                    c.value = val
                    c.number_format = self.money_fmt
                    c.alignment = self.cell_align
                elif fmt == "rate":
                    if val and val > 0:
                        c.value = val
                        c.number_format = self.rate_fmt
                    c.alignment = self.center_align
                elif fmt == "center":
                    c.value = val
                    c.alignment = self.center_align
                else:
                    c.value = val
                    c.alignment = self.cell_align

        # -------- 汇总行 --------
        total_r = header_row + 1 + len(detail)
        lt = f"合计 {len(detail)} 期"
        ws.cell(total_r, 1, lt).font = self.total_font
        ws.cell(total_r, 1).border = self.thin_border
        ws.cell(total_r, 1).alignment = self.center_align

        total_cols = [2, 6, 7, 8, 10, 11, 12, 14, 15, 16, 18, 19]
        for tc in total_cols:
            vals = []
            for item in detail:
                key_map = {2: "principal_balance", 6: "principal_due", 7: "principal_repaid",
                           8: "cumulative_overdue_principal", 10: "new_interest", 11: "repaid_interest",
                           12: "interest_balance", 14: "new_penalty", 15: "repaid_penalty",
                           16: "penalty_balance", 18: "new_interest_compound", 19: "interest_compound_balance"}
                key = key_map.get(tc)
                if key:
                    vals.append(item.get(key, 0))
            total_val = sum(vals)
            c = self._apply_money(ws, total_r, tc, round(total_val, 2))
            c.font = self.total_font

        # Column widths
        widths = [6, 14, 12, 12, 6, 12, 12, 14, 12, 14, 12, 14, 12, 14, 12, 14, 12, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # =================================================================
    # Sheet 3: 计算说明（公式核对）
    # =================================================================
    def _create_formula_sheet(self, wb, detail):
        ws = wb.create_sheet("计算说明")
        ws.sheet_properties.tabColor = "70AD47"

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"].value = "债权计算过程说明"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_align

        # Headers
        headers = ["期次", "计息区间", "利息计算式", "复利计算式", "罚息计算式",
                    "本金余额算式", "利息余额算式", "罚息余额算式", "复利余额算式", "备注"]
        self._apply_header(ws, headers, 3)

        for i, item in enumerate(detail):
            r = 4 + i
            ws.cell(r, 1, item.get("seq_no")).border = self.thin_border
            ws.cell(r, 1).alignment = self.center_align

            period = item.get("period", "")
            if not period:
                period = f"{item.get('start_date', '')} 至 {item.get('end_date', '')}"
            ws.cell(r, 2, period).border = self.thin_border
            ws.cell(r, 2).alignment = self.center_align

            ws.cell(r, 3, item.get("new_interest_formula", "")).border = self.thin_border
            ws.cell(r, 3).alignment = self.left_align
            ws.cell(r, 4, item.get("new_interest_compound_formula", "")).border = self.thin_border
            ws.cell(r, 4).alignment = self.left_align
            ws.cell(r, 5, item.get("new_penalty_formula", "")).border = self.thin_border
            ws.cell(r, 5).alignment = self.left_align
            ws.cell(r, 6, item.get("principal_balance_formula", "")).border = self.thin_border
            ws.cell(r, 6).alignment = self.left_align
            ws.cell(r, 7, item.get("interest_balance_formula", "")).border = self.thin_border
            ws.cell(r, 7).alignment = self.left_align
            ws.cell(r, 8, item.get("penalty_balance_formula", "")).border = self.thin_border
            ws.cell(r, 8).alignment = self.left_align
            ws.cell(r, 9, item.get("repayment_formula", "")).border = self.thin_border
            ws.cell(r, 9).alignment = self.left_align
            ws.cell(r, 10, item.get("remark", "")).border = self.thin_border
            ws.cell(r, 10).alignment = self.left_align

        widths = [6, 24, 40, 40, 40, 30, 30, 30, 30, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # =================================================================
    # Sheet 4: 还款记录
    # =================================================================
    def _create_records_sheet(self, wb, repayment_records, processed_details):
        ws = wb.create_sheet("还款记录")
        ws.sheet_properties.tabColor = "7030A0"

        ws.merge_cells("A1:I1")
        ws["A1"].value = "还款流水记录"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_align

        ws.cell(2, 1, f"共 {len(repayment_records)} 笔还款记录").font = Font(size=9, italic=True)
        ws.merge_cells("A2:I2")

        headers = ["序号", "还款日期", "还款类型", "还款金额", "归属周期",
                    "归属说明", "本金部分", "利息部分", "罚息部分", "复利部分"]
        self._apply_header(ws, headers, 4)

        for i, rec in enumerate(repayment_records):
            r = 5 + i
            ws.cell(r, 1, i + 1).border = self.thin_border; ws.cell(r, 1).alignment = self.center_align
            ws.cell(r, 2, str(rec.get("date", ""))[:10]).border = self.thin_border; ws.cell(r, 2).alignment = self.center_align

            rtype = rec.get("type", "")
            type_display = {"interest": "归还利息", "principal": "归还本金", "penalty": "归还罚息",
                            "compound": "归还复利", "fee": "支付费用", "利息": "归还利息", "本金": "归还本金",
                            "罚息": "归还罚息", "复利": "归还复利", "费用": "支付费用"}
            ws.cell(r, 3, type_display.get(rtype, rtype)).border = self.thin_border
            ws.cell(r, 3).alignment = self.center_align

            self._apply_money(ws, r, 4, rec.get("amount", 0))

            # Find matching processed detail
            detail = None
            for pd_item in processed_details:
                if str(pd_item.get("date", ""))[:10] == str(rec.get("date", ""))[:10]:
                    detail = pd_item
                    break

            ws.cell(r, 5, detail.get("period", "") if detail else "").border = self.thin_border
            ws.cell(r, 5).alignment = self.center_align
            ws.cell(r, 6, detail.get("remark", "") if detail else "").border = self.thin_border
            ws.cell(r, 6).alignment = self.center_align

            alloc = detail.get("allocation", {}) if detail else {}
            self._apply_money(ws, r, 7, alloc.get("principal", 0) if alloc else 0)
            self._apply_money(ws, r, 8, alloc.get("interest", 0) if alloc else 0)
            self._apply_money(ws, r, 9, alloc.get("penalty", 0) if alloc else 0)
            self._apply_money(ws, r, 10, alloc.get("compound", 0) if alloc else 0)

        widths = [6, 14, 12, 14, 26, 16, 14, 14, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # =================================================================
    # 判决模式导出
    # =================================================================
    def export_judgment_result(self, r: Dict[str, Any]) -> bytes:
        """导出判决债权统计Excel（3个Sheet）。"""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        jdg_date = r.get("judgment_date", "")
        cutoff = r.get("cutoff", "")
        eff_date = r.get("effective_date", "")
        perf_days = r.get("performance_days", 10)
        delay_days = r.get("delay_days", 0)

        # ===== Sheet 1: 判决债权汇总 =====
        ws1 = wb.create_sheet("判决债权汇总")
        ws1.sheet_properties.tabColor = "4472C4"

        ws1.merge_cells("A1:D1")
        ws1["A1"].value = "最新债权统计表（判决）"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1["A1"].alignment = Alignment(horizontal="center")

        ws1["A2"] = f"判决日：{jdg_date}    截算日：{cutoff}    计息天数：{r.get('total_days', 0)}天"
        ws1.merge_cells("A2:D2")

        if eff_date:
            ws1["A3"] = f"生效日：{eff_date}    履行期：{perf_days}天    迟延履行利息天数：{delay_days}天"
            ws1.merge_cells("A3:D3")

        header_row = 5
        headers = ["项目", "判决确认金额", "最新统计金额", "备注"]
        for i, h in enumerate(headers, 1):
            c = ws1.cell(header_row, i, h)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.alignment = Alignment(horizontal="center")
            c.border = self.thin_border

        items = [
            ("本金", r.get("principal_original", 0), r.get("total_principal", 0), ""),
            ("利息", r.get("interest_original", 0), r.get("total_interest", 0), f"判决后新增{r.get('interest_new',0):,.2f}"),
            ("罚息", r.get("penalty_original", 0), r.get("total_penalty", 0), f"判决后新增{r.get('penalty_new',0):,.2f}"),
            ("复利", r.get("compound_original", 0), r.get("total_compound", 0), f"判决后新增{r.get('compound_new',0):,.2f}"),
        ]
        lf = r.get("lawyer_fee", 0)
        cf = r.get("court_fee", 0)
        pf = r.get("preservation_fee", 0)
        if lf > 0: items.append(("律师费", lf, lf, ""))
        if cf > 0: items.append(("案件受理费", cf, cf, ""))
        if pf > 0: items.append(("保全费", pf, pf, ""))
        di = r.get("delay_interest", 0)
        if di > 0:
            items.append(("迟延履行利息", "—", di, f"日万分之一点七五×{delay_days}天"))
        items.append(("", "", "", ""))
        items.append(("债权总额", "", r.get("grand_total", 0), ""))

        for i, (label, jdg_amt, cur_amt, note) in enumerate(items):
            row = header_row + 1 + i
            ws1.cell(row, 1, label).border = self.thin_border
            ws1.cell(row, 1).alignment = self.center_align
            if label == "债权总额":
                ws1.cell(row, 1).font = Font(bold=True, size=10, color="FF0000")

            if isinstance(jdg_amt, (int, float)) and jdg_amt > 0:
                c = ws1.cell(row, 2, jdg_amt)
                c.number_format = self.money_fmt
            else:
                ws1.cell(row, 2, jdg_amt if jdg_amt else 0)
            ws1.cell(row, 2).border = self.thin_border
            ws1.cell(row, 2).alignment = self.cell_align

            c = ws1.cell(row, 3, cur_amt)
            if isinstance(cur_amt, (int, float)) and cur_amt > 0:
                c.number_format = self.money_fmt
            c.border = self.thin_border
            c.alignment = self.cell_align
            if label == "债权总额":
                c.font = Font(bold=True, size=11, color="FF0000")

            ws1.cell(row, 4, note).border = self.thin_border
            ws1.cell(row, 4).alignment = self.left_align

        ws1.column_dimensions["A"].width = 16
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 20
        ws1.column_dimensions["D"].width = 30

        # ===== Sheet 2: 计算明细 =====
        ws2 = wb.create_sheet("计算明细")
        ws2.sheet_properties.tabColor = "E06C00"

        ws2.merge_cells("A1:E1")
        ws2["A1"].value = f"计算明细（判决日{jdg_date} → 截算日{cutoff}，{r.get('total_days',0)}天，365日/年）"
        ws2["A1"].font = Font(bold=True, size=12)

        h2 = ["项目", "基数", "年利率", "天数", "金额"]
        for i, h in enumerate(h2, 1):
            c = ws2.cell(3, i, h)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.alignment = Alignment(horizontal="center")
            c.border = self.thin_border

        detail_rows = [
            ("判决确认本金（新增利息基数）", r.get("principal_original", 0), r.get("post_rate", 0),
             r.get("total_days", 0), r.get("interest_new", 0),
             f"{r.get('principal_original',0):,.2f} × {r.get('post_rate',0):.2f}% × {r.get('total_days',0)} / 365"),
            ("判决确认本金（新增罚息基数）", r.get("principal_original", 0),
             r.get("post_rate", 0) * 1.5,
             r.get("total_days", 0), r.get("penalty_new", 0),
             f"{r.get('principal_original',0):,.2f} × {r.get('post_rate',0)*1.5:.2f}% × {r.get('total_days',0)} / 365"),
            ("欠付利息余额（新增复利基数）",
             r.get("interest_original", 0) + r.get("interest_new", 0),
             r.get("post_rate", 0),
             r.get("total_days", 0), r.get("compound_new", 0),
             f"利息余额 × {r.get('post_rate',0):.2f}% × {r.get('total_days',0)} / 365"),
        ]
        if r.get("delay_interest", 0) > 0:
            detail_rows.append((
                "迟延履行利息", r.get("principal_original", 0),
                0.0175 * 365 / 100, r.get("delay_days", 0), r.get("delay_interest", 0),
                f"{r.get('principal_original',0):,.2f} × 0.0175%/日 × {r.get('delay_days',0)}天"
            ))

        for i, (label, base, rate, days, amount, formula) in enumerate(detail_rows):
            row = 4 + i
            ws2.cell(row, 1, label).border = self.thin_border
            ws2.cell(row, 1).alignment = self.left_align
            c = ws2.cell(row, 2, round(base, 2)); c.number_format = self.money_fmt; c.border = self.thin_border; c.alignment = self.cell_align
            ws2.cell(row, 3, round(rate, 4)).border = self.thin_border; ws2.cell(row, 3).alignment = self.center_align
            ws2.cell(row, 4, days).border = self.thin_border; ws2.cell(row, 4).alignment = self.center_align
            c = ws2.cell(row, 5, round(amount, 2)); c.number_format = self.money_fmt; c.border = self.thin_border; c.alignment = self.cell_align
            ws2.cell(row, 6, formula).border = self.thin_border; ws2.cell(row, 6).alignment = self.left_align

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 8
        ws2.column_dimensions["E"].width = 16
        ws2.column_dimensions["F"].width = 50

        # ===== Sheet 3: 计算过程 =====
        ws3 = wb.create_sheet("计算过程")
        ws3.sheet_properties.tabColor = "70AD47"

        ws3.merge_cells("A1:C1")
        ws3["A1"].value = "计算过程说明"
        ws3["A1"].font = Font(bold=True, size=12)

        calc_steps = [
            ("一、基础信息", "", ""),
            ("判决日", jdg_date, "法院判决作出的日期"),
            ("判决确认本金", f'{r.get("principal_original",0):,.2f}', "法院判决确认的尚欠本金余额"),
            ("判决确认利息", f'{r.get("interest_original",0):,.2f}', "法院判决确认的欠付利息"),
            ("判决确认罚息", f'{r.get("penalty_original",0):,.2f}', "法院判决确认的欠付罚息"),
            ("判决确认复利", f'{r.get("compound_original",0):,.2f}', "法院判决确认的欠付复利"),
            ("判决后年利率", f'{r.get("post_rate",0):.2f}%', "按照判决确定的利率"),
            ("计息基数", "365日/年", "判决后按实际日历天数"),
            ("截算日", cutoff, "最新欠款日"),
            ("", "", ""),
            ("二、新增利息", "", ""),
            ("公式", f'本金 × 年利率 × 天数 / 365', ""),
            ("计算", f'{r.get("principal_original",0):,.2f} × {r.get("post_rate",0):.2f}% × {r.get("total_days",0)} / 365', ""),
            ("结果", f'{r.get("interest_new",0):,.2f}', ""),
            ("", "", ""),
            ("三、新增罚息", "", ""),
            ("公式", f'本金 × (年利率×1.5) × 天数 / 365', ""),
            ("计算", f'{r.get("principal_original",0):,.2f} × {r.get("post_rate",0)*1.5:.2f}% × {r.get("total_days",0)} / 365', ""),
            ("结果", f'{r.get("penalty_new",0):,.2f}', ""),
            ("", "", ""),
            ("四、新增复利", "", ""),
            ("公式", f'(判决利息 + 新增利息) × 年利率 × 天数 / 365', ""),
            ("计算", f'({r.get("interest_original",0):,.2f} + {r.get("interest_new",0):,.2f}) × {r.get("post_rate",0):.2f}% × {r.get("total_days",0)} / 365', ""),
            ("结果", f'{r.get("compound_new",0):,.2f}', ""),
        ]
        if r.get("delay_interest", 0) > 0:
            delay_start_str = f'判决生效日{r.get("effective_date","")} + {r.get("performance_days",10)}天履行期'
            calc_steps += [
                ("", "", ""),
                ("五、迟延履行利息", "", ""),
                ("依据", "《民事诉讼法》第264条", "日万分之一点七五"),
                ("起算", delay_start_str, "生效日+履行期届满次日起"),
                ("公式", "本金 × 0.0175%/日 × 天数", ""),
                ("计算", f'{r.get("principal_original",0):,.2f} × 0.0175% × {r.get("delay_days",0)}天', ""),
                ("结果", f'{r.get("delay_interest",0):,.2f}', ""),
            ]

        for i, (step, val, note) in enumerate(calc_steps):
            row = 3 + i
            ws3.cell(row, 1, step).border = self.thin_border
            ws3.cell(row, 1).alignment = self.left_align
            ws3.cell(row, 2, val).border = self.thin_border
            ws3.cell(row, 2).alignment = self.left_align
            ws3.cell(row, 3, note).border = self.thin_border
            ws3.cell(row, 3).alignment = self.left_align

        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 50
        ws3.column_dimensions["C"].width = 36

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    # =================================================================
    # 原贷款流水模板导出（保留）
    # =================================================================
    def _export_loan_flow_template_result(self, calc_result, loan_data, repay_records):
        """原流水模板导出（NHRCB格式，保留兼容）"""
        wb = Workbook()
        ws = wb.active
        ws.title = self._build_flow_sheet_name(loan_data)
        detail = calc_result.get("detail", [])
        summary = calc_result.get("summary", {})

        calc_date = loan_data.get("calculation_date") or summary.get("calculation_date", "")
        title = f"本息计算表（暂计至{calc_date}）" if calc_date else "本息计算表"
        ws["A1"] = title
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:S1")

        accel_date = loan_data.get("acceleration_date", "")
        ws["A2"] = f"贷款金额：{loan_data.get('loan_amount', '')}元"
        ws["A3"] = f"借款期限：{loan_data.get('start_date', '')}至{loan_data.get('end_date', '')}"
        ws["A4"] = f"贷款期限以提前到期日为到期日，即{accel_date}" if accel_date else ""
        ws.merge_cells("A2:S2"); ws.merge_cells("A3:S3"); ws.merge_cells("A4:S4")

        headers = [
            "序号", "本金余额", "起算日", "暂计至", "天数", "应还本金", "已还本金", "累计逾期本金",
            "正常贷款年利率", "应付利息", "已还利息", "累计尚欠利息",
            "罚息年利率", "应付罚息", "已还罚息", "累计尚欠罚息",
            "复利年利率", "应付复利", "累计尚欠复利",
        ]
        self._apply_header(ws, headers, 7)

        for i, item in enumerate(detail):
            r = 8 + i
            vals = [
                item.get("seq_no"), item.get("principal_balance"),
                item.get("start_date"), item.get("end_date"), item.get("days"),
                item.get("principal_due", 0), item.get("principal_repaid", 0),
                item.get("cumulative_overdue_principal", 0),
                None, item.get("new_interest", 0), item.get("repaid_interest", 0),
                item.get("interest_balance", 0),
                None, item.get("new_penalty", 0), item.get("repaid_penalty", 0),
                item.get("penalty_balance", 0),
                None, item.get("new_interest_compound", 0),
                item.get("interest_compound_balance", 0),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = self.thin_border
                if c in (9, 12, 16):
                    cell.alignment = self.center_align
                elif c in (3, 4):
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.cell_align
                    if isinstance(v, (int, float)):
                        cell.number_format = self.money_fmt

        if self.auto_width if hasattr(self, 'auto_width') else True:
            widths = [8, 14, 12, 12, 8, 12, 12, 14, 12, 12, 12, 14, 12, 12, 12, 14, 12, 12, 14]
            for idx, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + idx)].width = w

        # Formula sheet
        self._create_formula_sheet(wb, detail)

        if repay_records:
            self._create_records_sheet(wb, repay_records, calc_result.get("repayment_details", []))

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    def _build_flow_sheet_name(self, loan_data):
        amount = loan_data.get("loan_amount", 0)
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            amt = 0
        return f"{int(round(amt / 10000))}万" if amt >= 10000 else "贷款流水"
