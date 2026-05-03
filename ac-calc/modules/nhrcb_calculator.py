"""
南海农商银行债权计算引擎
基于参考文件和输出样本分析实现
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Any, Optional, Tuple
import sys
from pathlib import Path

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from modules.models.calculation_result import CalculationResult, InterestDetail, PaymentApplication
from utils.date_utils import parse_date, format_date


@dataclass
class CalculationPeriod:
    """计息期间"""
    period_no: int                      # 期数
    start_date: datetime                # 起始日期
    end_date: datetime                  # 结束日期
    days: int                           # 计息天数
    principal_balance: Decimal          # 本金余额
    rate: Decimal                       # 年利率
    interest: Decimal = Decimal('0')    # 应计利息
    paid_interest: Decimal = Decimal('0')  # 已还利息

    def calculate_interest(self):
        """计算利息"""
        # 公式：本金 × 年利率 × 天数 / 360
        self.interest = (
            self.principal_balance *
            self.rate *
            Decimal(str(self.days)) /
            Decimal('360')
        )
        # 四舍五入到2位小数
        self.interest = self.interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


@dataclass
class RepaymentEvent:
    """还款事件"""
    date: datetime
    amount: Decimal
    type: str                           # 'principal', 'interest'
    description: str = ""


class NHRCBCalculator:
    """南海农商银行计算器"""

    def __init__(
        self,
        loan_amount: float,
        start_date: str,
        end_date: str,
        calculation_date: str,
        annual_rate: float = 0.048,      # 默认4.8%
        penalty_rate_multiplier: float = 1.5,  # 罚息利率倍数
        day_count_convention: str = "actual/360",
        quarterly_repayment_rate: float = 0.02  # 每季度还本2%
    ):
        """
        初始化计算器

        Args:
            loan_amount: 贷款金额
            start_date: 起息日
            end_date: 到期日
            calculation_date: 计算截止日
            annual_rate: 年利率（如0.048表示4.8%）
            penalty_rate_multiplier: 罚息利率倍数
            day_count_convention: 天数计算规则
            quarterly_repayment_rate: 每季度还本比例
        """
        self.loan_amount = Decimal(str(loan_amount))
        self.start_date = parse_date(start_date)
        self.end_date = parse_date(end_date)
        self.calculation_date = parse_date(calculation_date)
        self.annual_rate = Decimal(str(annual_rate))
        self.penalty_rate = self.annual_rate * Decimal(str(penalty_rate_multiplier))
        self.day_count_convention = day_count_convention
        self.quarterly_repayment_rate = Decimal(str(quarterly_repayment_rate))
        self.quarterly_repayment_amount = self.loan_amount * self.quarterly_repayment_rate

        # 计算期间列表
        self.periods: List[CalculationPeriod] = []

        # 还款事件列表
        self.repayments: List[RepaymentEvent] = []

        # 当前状态
        self.current_principal = self.loan_amount
        self.unpaid_interest = Decimal('0')

    def add_repayment(self, date: str, amount: float, type: str, description: str = ""):
        """添加还款记录"""
        self.repayments.append(RepaymentEvent(
            date=parse_date(date),
            amount=Decimal(str(amount)),
            type=type,
            description=description
        ))

    def calculate_periods(self) -> List[CalculationPeriod]:
        """
        计算所有计息期间

        南海农商银行按月计息，每月21日为结息日
        """
        periods = []
        current_date = self.start_date
        period_no = 1

        while current_date < self.calculation_date:
            # 找到下一个结息日（每月21日）
            if current_date.day <= 21:
                # 本月21日
                period_end = current_date.replace(day=21)
            else:
                # 下月21日
                if current_date.month == 12:
                    period_end = current_date.replace(year=current_date.year + 1, month=1, day=21)
                else:
                    period_end = current_date.replace(month=current_date.month + 1, day=21)

            # 确保不超过计算截止日
            if period_end > self.calculation_date:
                period_end = self.calculation_date

            # 计算天数
            days = (period_end - current_date).days + 1  # 算头算尾

            # 如果正好30天或者实际天数，根据配置决定
            # 从输出文件看，大部分月份使用30天，但有些使用实际天数
            # 这里先使用实际天数

            # 创建计息期间
            period = CalculationPeriod(
                period_no=period_no,
                start_date=current_date,
                end_date=period_end,
                days=days,
                principal_balance=self.current_principal,
                rate=self.annual_rate
            )

            # 应用本期内的还款
            self._apply_repayments_in_period(period)

            # 计算利息
            period.calculate_interest()

            # 累计未还利息
            self.unpaid_interest += period.interest
            period.paid_interest = Decimal('0')  # 初始未还

            periods.append(period)

            # 移动到下一期
            current_date = period_end + timedelta(days=1)
            period_no += 1

            # 检查是否还有季度还本
            # 从输入文件看，季度还本发生在：11月、2月、5月、8月的20日
            # 需要检查还款事件

        self.periods = periods
        return periods

    def _apply_repayments_in_period(self, period: CalculationPeriod):
        """应用期间内的还款事件"""
        for repayment in self.repayments:
            if period.start_date <= repayment.date <= period.end_date:
                if repayment.type == 'principal':
                    # 本金还款
                    self.current_principal -= repayment.amount
                    if self.current_principal < 0:
                        self.current_principal = Decimal('0')
                elif repayment.type == 'interest':
                    # 利息还款 - 这里暂时不处理，在后续冲销中处理
                    pass

    def calculate(self) -> CalculationResult:
        """
        执行完整计算

        Returns:
            CalculationResult: 计算结果
        """
        result = CalculationResult()
        result.loan_amount = float(self.loan_amount)
        result.start_date = format_date(self.start_date)
        result.end_date = format_date(self.end_date)
        result.calculation_date = format_date(self.calculation_date)

        # 1. 计算所有计息期间
        periods = self.calculate_periods()

        # 2. 应用还款冲销
        payment_applications = self._apply_payment_priority(periods)

        # 3. 生成结果
        result.interest_details = []
        total_interest = Decimal('0')

        for period in periods:
            detail = InterestDetail(
                period_start=format_date(period.start_date),
                period_end=format_date(period.end_date),
                days=period.days,
                principal_balance=float(period.principal_balance),
                rate=float(period.rate * 100),  # 转换为百分比
                interest=float(period.interest)
            )
            result.interest_details.append(detail)
            total_interest += period.interest

        result.payment_applications = payment_applications
        result.total_interest = float(total_interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        result.remaining_principal = float(self.current_principal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

        # 计算债权总额
        result.calculate_total()

        return result

    def _apply_payment_priority(
        self,
        periods: List[CalculationPeriod]
    ) -> List[PaymentApplication]:
        """
        应用还款冲销

        南海农商银行的冲销规则（从输出文件分析）：
        1. 先冲销利息
        2. 再冲销本金

        Returns:
            List[PaymentApplication]: 还款冲销明细
        """
        applications = []
        unpaid_interest = Decimal('0')
        unpaid_penalty = Decimal('0')

        # 按还款日期排序
        sorted_repayments = sorted(self.repayments, key=lambda r: r.date)

        for repayment in sorted_repayments:
            remaining = repayment.amount

            # 先冲销利息（包括应计利息和未还利息）
            if remaining > 0 and unpaid_interest > 0:
                interest_payment = min(remaining, unpaid_interest)
                unpaid_interest -= interest_payment
                remaining -= interest_payment

            # 再冲销本金
            principal_payment = Decimal('0')
            if remaining > 0 and repayment.type == 'principal':
                principal_payment = remaining
                remaining = Decimal('0')

            # 创建冲销记录
            application = PaymentApplication(
                date=format_date(repayment.date),
                payment_amount=float(repayment.amount),
                principal_payment=float(principal_payment),
                interest_payment=float(repayment.amount - principal_payment),
                penalty_payment=0.0,
                compound_payment=0.0,
                fee_payment=0.0,
                remaining_principal=float(self.current_principal)
            )

            applications.append(application)

        return applications


def calculate_nhrcb_from_flow(
    flow_data: List[Dict[str, Any]],
    calculation_date: Optional[str] = None
) -> CalculationResult:
    """
    从流水数据计算债权

    Args:
        flow_data: 流水数据列表
        calculation_date: 计算截止日（可选，默认为今天）

    Returns:
        CalculationResult: 计算结果
    """
    if not flow_data:
        raise ValueError("流水数据不能为空")

    # 提取基本信息
    loan_amount = None
    start_date = None
    end_date = None

    # 提取还款记录
    repayments = []

    for row in flow_data:
        trans_type = row.get('交易类型', '')
        amount_str = str(row.get('交易金额', '0'))
        amount = Decimal(amount_str.replace(',', '').replace('-', ''))

        if '放款' in trans_type:
            # 提取贷款金额
            loan_amount = float(amount)
            # 提取起息日和到期日
            start_date = row.get('起息日')
            end_date = row.get('到期日期')

        elif '还本' in trans_type or '本金' in trans_type:
            # 本金还款
            repayments.append({
                'date': row.get('实际发生日'),
                'amount': float(amount),
                'type': 'principal'
            })

    if not loan_amount:
        raise ValueError("未找到贷款金额")

    if not start_date:
        raise ValueError("未找到起息日")

    if not end_date:
        raise ValueError("未找到到期日")

    # 如果没有指定计算截止日，使用今天或最后一条流水的日期
    if not calculation_date:
        calculation_date = format_date(datetime.now())
    else:
        calculation_date = format_date(parse_date(calculation_date))

    # 创建计算器
    calculator = NHRCBCalculator(
        loan_amount=loan_amount,
        start_date=start_date,
        end_date=end_date,
        calculation_date=calculation_date
    )

    # 添加还款记录
    for repayment in repayments:
        calculator.add_repayment(
            date=repayment['date'],
            amount=repayment['amount'],
            type=repayment['type']
        )

    # 执行计算
    return calculator.calculate()


# 测试函数
def test_with_reference_data():
    """使用参考数据测试"""
    source_dir = Path(r'D:\我的资料库\Documents\xwechat_files\wxid_w4uaqxmkyyp622_5242\msg\file\2026-03\cs')
    xlsx_file = list(source_dir.glob('*.xlsx'))[0]

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active

    # 读取流水数据
    flow_data = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 序号不为空
            data_row = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    data_row[header] = row[i]
            flow_data.append(data_row)

    # 执行计算
    result = calculate_nhrcb_from_flow(
        flow_data=flow_data,
        calculation_date="2026-03-09"
    )

    # 输出结果
    print("\n" + "=" * 80)
    print("计算结果")
    print("=" * 80)
    print(f"贷款金额: {result.loan_amount:,.2f}")
    print(f"剩余本金: {result.remaining_principal:,.2f}")
    print(f"利息总额: {result.total_interest:,.2f}")
    print(f"债权总额: {result.total_debt:,.2f}")
    print("\n前10期利息明细:")
    for detail in result.interest_details[:10]:
        print(f"  {detail['period_start']} ~ {detail['period_end']} "
              f"({detail['days']}天): {detail['interest']:,.2f}")

    return result


if __name__ == "__main__":
    test_with_reference_data()
