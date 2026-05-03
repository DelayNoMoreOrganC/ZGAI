"""
南海农商银行Excel导出模块
按照参考格式输出债权计算表
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from typing import List


class NHRCBExcelExporter:
    """南海农商银行Excel导出器"""

    def __init__(self):
        self.wb = None
        self.ws = None

    def export(self, result, loan_info: dict) -> BytesIO:
        """
        导出为Excel文件

        Args:
            result: CalculationResult对象
            loan_info: 贷款信息字典

        Returns:
            BytesIO: Excel文件内容
        """
        # 创建工作簿
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"{int(loan_info.get('loan_amount', 0))}万"

        # 写入表头信息
        self._write_header(loan_info)

        # 写入数据列标题
        self._write_column_headers()

        # 写入计息明细
        self._write_interest_details(result)

        # 写入汇总信息
        self._write_summary(result, loan_info)

        # 设置列宽
        self._set_column_widths()

        # 保存到BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def _write_header(self, loan_info: dict):
        """写入表头信息"""
        ws = self.ws

        # 第1行：标题
        calculation_date = loan_info.get('calculation_date', datetime.now().strftime('%Y-%m-%d'))
        year = calculation_date[:4]
        month = calculation_date[5:7]
        day = calculation_date[8:10]

        ws['A1'] = f'债权计算信息汇总表（截至{year}年{month}月{day}日）'
        ws['A1'].font = Font(name='宋体', size=14, bold=True)
        ws.merge_cells('A1:T1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 第2行：客户信息（示例）
        ws['A2'] = '借 款 人：广东南海农村商业银行股份有限公司里水支行                                                       \n账 户：佛山市、南海区、顺德区......有限公司'
        ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A2:T2')

        # 第3行：合同信息
        loan_amount = loan_info.get('loan_amount', 0)
        start_date = loan_info.get('start_date', '')
        end_date = loan_info.get('end_date', '')

        ws['A3'] = f'合同编号：与客户编号一致（2023年0320号）《综合授信合同》\n贷款金额：{loan_amount}元\n起止期限：{start_date}至{end_date}\n执行利率：按放款日全国银行间同业拆借中心公布的贷款市场报价利率（LPR）125基点形成，即每年1月1日调整一次。\n还款方式：按月计息，到期还本。资金发放后每季度归还本金2%，（贷款到期转逾期）。\n其他情况：已还款清偿完毕，视为结清；现计算日期：{year}年{month}月{day}日。'
        ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A3:T3')

        # 设置行高
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 90

    def _write_column_headers(self):
        """写入数据列标题"""
        ws = self.ws

        # 列标题
        headers = [
            '期数', '本金余额', '起息日', '止息日', '天数',
            '本期本金', '累计归还本金', '剩余本金',
            '正常利率', '应计利息', '已还利息', '累计未还利息',
            '罚息利率', '应计罚息', '已还罚息', '累计未还罚息',
            '复利利率', '应计复利', '累计未还复利'
        ]

        # 从第5行开始（前4行是表头信息）
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(name='宋体', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # 设置行高
        ws.row_dimensions[5].height = 25

    def _write_interest_details(self, result):
        """写入计息明细"""
        ws = self.ws

        # 从第6行开始写入数据
        row = 6
        current_principal = result.loan_amount
        total_repaid_principal = 0
        total_unpaid_interest = 0

        for i, detail in enumerate(result.interest_details):
            # 期数
            ws.cell(row=row, column=1, value=i + 1)

            # 本金余额
            ws.cell(row=row, column=2, value=current_principal)

            # 起息日（转换为Excel日期）
            start_date = datetime.strptime(detail.period_start, '%Y-%m-%d')
            start_excel = self._date_to_excel(start_date)
            ws.cell(row=row, column=3, value=start_excel)

            # 止息日
            end_date = datetime.strptime(detail.period_end, '%Y-%m-%d')
            end_excel = self._date_to_excel(end_date)
            ws.cell(row=row, column=4, value=end_excel)

            # 天数
            ws.cell(row=row, column=5, value=detail.days)

            # 本期本金（0，因为没有每期还款）
            ws.cell(row=row, column=6, value=0)

            # 累计归还本金（0）
            ws.cell(row=row, column=7, value=0)

            # 剩余本金
            ws.cell(row=row, column=8, value=current_principal)

            # 正常利率
            ws.cell(row=row, column=9, value=detail.rate / 100)

            # 应计利息
            ws.cell(row=row, column=10, value=detail.interest)

            # 已还利息（0）
            ws.cell(row=row, column=11, value=0)

            # 累计未还利息
            total_unpaid_interest += detail.interest
            ws.cell(row=row, column=12, value=total_unpaid_interest)

            # 罚息利率
            penalty_rate = 0.072
            ws.cell(row=row, column=13, value=penalty_rate)

            # 应计罚息（0，因为没有逾期）
            ws.cell(row=row, column=14, value=0)

            # 已还罚息（0）
            ws.cell(row=row, column=15, value=0)

            # 累计未还罚息（0）
            ws.cell(row=row, column=16, value=0)

            # 复利利率
            ws.cell(row=row, column=17, value=penalty_rate)

            # 应计复利（0）
            ws.cell(row=row, column=18, value=0)

            # 累计未还复利（0）
            ws.cell(row=row, column=19, value=0)

            # 设置对齐方式
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col in [1]:  # 期数居中
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

    def _write_summary(self, result, loan_info: dict):
        """写入汇总信息"""
        ws = self.ws

        # 在最后一行后添加汇总
        summary_row = ws.max_row + 2

        # 债权汇总
        ws.cell(row=summary_row, column=1, value='债权汇总')
        ws.cell(row=summary_row, column=1).font = Font(name='宋体', size=12, bold=True)
        ws.merge_cells(f'A{summary_row}:B{summary_row}')

        # 剩余本金
        ws.cell(row=summary_row + 1, column=1, value='剩余本金：')
        ws.cell(row=summary_row + 1, column=2, value=result.remaining_principal)

        # 利息总额
        ws.cell(row=summary_row + 2, column=1, value='利息总额：')
        ws.cell(row=summary_row + 2, column=2, value=result.total_interest)

        # 罚息总额
        if result.total_penalty > 0:
            ws.cell(row=summary_row + 3, column=1, value='罚息总额：')
            ws.cell(row=summary_row + 3, column=2, value=result.total_penalty)

        # 复利总额
        if result.total_compound > 0:
            ws.cell(row=summary_row + 4, column=1, value='复利总额：')
            ws.cell(row=summary_row + 4, column=2, value=result.total_compound)

        # 债权总额
        total_row = summary_row + 5
        ws.cell(row=total_row, column=1, value='债权总额：')
        ws.cell(row=total_row, column=1).font = Font(name='宋体', size=12, bold=True)
        ws.cell(row=total_row, column=2, value=result.total_debt)
        ws.cell(row=total_row, column=2).font = Font(name='宋体', size=12, bold=True, color='FF0000')

    def _set_column_widths(self):
        """设置列宽"""
        ws = self.ws

        # 设置列宽
        column_widths = {
            'A': 8,   # 期数
            'B': 15,  # 本金余额
            'C': 12,  # 起息日
            'D': 12,  # 止息日
            'E': 10,  # 天数
            'F': 12,  # 本期本金
            'G': 15,  # 累计归还本金
            'H': 15,  # 剩余本金
            'I': 12,  # 正常利率
            'J': 15,  # 应计利息
            'K': 15,  # 已还利息
            'L': 15,  # 累计未还利息
            'M': 12,  # 罚息利率
            'N': 15,  # 应计罚息
            'O': 15,  # 已还罚息
            'P': 15,  # 累计未还罚息
            'Q': 12,  # 复利利率
            'R': 15,  # 应计复利
            'S': 15,  # 累计未还复利
            'T': 15   # 预留
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _date_to_excel(self, date: datetime) -> int:
        """
        将datetime转换为Excel日期数字

        Args:
            date: datetime对象

        Returns:
            int: Excel日期数字
        """
        # Excel日期基准：1900年1月1日 = 1
        delta = date - datetime(1899, 12, 30)
        return delta.days


def export_calculation_result(result, loan_info: dict) -> BytesIO:
    """
    导出计算结果为Excel

    Args:
        result: CalculationResult对象
        loan_info: 贷款信息字典

    Returns:
        BytesIO: Excel文件内容
    """
    exporter = NHRCBExcelExporter()
    return exporter.export(result, loan_info)
