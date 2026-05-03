"""
银行贷款流水模板解析与计算。

当前先落地“伍钧乐欠款流水”这一类贷款台账流水：
- 支持 PDF/OCR 文本与 xlsx 流水
- 将流水转换为标准事件
- 生成与参考 Excel 同口径的债权计算表
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from utils.date_utils import format_date, parse_date
from utils.constants import RepaymentType

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None


getcontext().prec = 28

TEMPLATE_KEY = "loan_flow_nanhai_rural_bank"
TEMPLATE_NAME = "贷款流水模板"
TEMPLATE_BANK_NAME = "银行贷款流水"

FLOW_KEYWORDS = (
    "存款账户转账放款",
    "正常本金结息",
    "逾期本金罚息结息",
    "现金回收",
    "未到期本金到期转桶子",
)


def maybe_extract_loan_flow_preview(
    file_bytes: Optional[bytes] = None,
    filename: str = "",
    text: str = "",
) -> Optional[Dict[str, Any]]:
    """尝试识别并提取银行贷款流水模板。"""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    source_kind = "ocr_text"

    if suffix in {"xlsx", "xls"} and file_bytes:
        rows, warnings = extract_flow_rows_from_spreadsheet(file_bytes, f".{suffix}")
        source_kind = "spreadsheet"
    elif text:
        if not is_loan_flow_text(text):
            return None
        rows = extract_flow_rows_from_text(text)
    else:
        return None

    selected_rows, selection_warnings = _select_relevant_loan_rows(rows, filename)
    warnings.extend(selection_warnings)

    events = classify_flow_events(selected_rows)
    if not events:
        return None

    has_principal_repayment = any(item["event_type"] == "principal_repayment" for item in events)
    has_due_transfer = any(item["event_type"] == "principal_due_transfer" for item in events)
    if has_principal_repayment and not has_due_transfer:
        warnings.append("当前流水未显式包含“应还本金转逾期”节点；如合同约定分期还本，请人工补充还本计划或到期节点。")

    loan_amount = _first_event_amount(events, "disbursement")
    start_date = _first_event_date(events, "disbursement") or _first_non_empty(selected_rows, "loan_start_date")
    end_date = _first_non_empty(selected_rows, "contract_end_date")
    calculation_date = _last_event_date(events)
    actual_principal_paid = sum(
        Decimal(str(item["amount"]))
        for item in events
        if item["event_type"] == "principal_repayment"
    )
    remaining_principal = None
    if loan_amount is not None:
        remaining_principal = float(max(Decimal(str(loan_amount)) - actual_principal_paid, Decimal("0")))

    repayment_records = []
    for event in events:
        repayment_type = _event_type_to_repayment_type(event["event_type"])
        if repayment_type is None:
            continue
        repayment_records.append(
            {
                "date": event["date"],
                "type": repayment_type,
                "amount": round(float(event["amount"]), 2),
                "description": event.get("description", ""),
                "source": event.get("raw_line", ""),
            }
        )

    field_sources: Dict[str, str] = {}
    if loan_amount is not None:
        disbursement = next((item for item in events if item["event_type"] == "disbursement"), None)
        field_sources["loan_amount"] = (disbursement or {}).get("raw_line", "流水放款记录")
    if start_date:
        field_sources["start_date"] = next(
            (item.get("raw_line", "") for item in events if item["event_type"] == "disbursement"),
            "流水放款记录",
        )
    if end_date:
        field_sources["end_date"] = next(
            (row.get("raw_line", "") for row in selected_rows if row.get("contract_end_date")),
            "流水到期日期",
        )
    if calculation_date:
        field_sources["calculation_date"] = next(
            (item.get("raw_line", "") for item in reversed(events)),
            "最后一笔流水日期",
        )
    if remaining_principal is not None:
        field_sources["remaining_principal"] = "贷款本金 - 已归还本金"
    if repayment_records:
        field_sources["repayment_records"] = "贷款流水中的现金回收记录"

    data = {
        "loan_amount": loan_amount,
        "annual_interest_rate": None,
        "start_date": start_date or "",
        "end_date": end_date or "",
        "calculation_date": calculation_date or "",
        "remaining_principal": remaining_principal,
        "accrued_interest": 0,
        "penalty_interest": 0,
        "compound_interest": 0,
        "acceleration_date": "",
        "_template_key": TEMPLATE_KEY,
        "_template_name": TEMPLATE_NAME,
        "_bank_name": TEMPLATE_BANK_NAME,
        "_template_payload": {
            "source_kind": source_kind,
            "events": events,
            "raw_rows": selected_rows,
        },
    }

    if source_kind == "spreadsheet":
        warnings.append("已按结构化流水解析；利率和提前到期日如无法从流水中直接识别，需人工填写。")
    else:
        warnings.append("当前为 OCR 解析结果，建议优先复核日期、金额与本金到期转逾期节点。")

    return {
        "template_key": TEMPLATE_KEY,
        "template_name": TEMPLATE_NAME,
        "data": data,
        "repayment_records": repayment_records,
        "pending_repayment_records": [],
        "field_sources": field_sources,
        "warnings": warnings,
        "raw_llm_data": {},
    }


def is_loan_flow_text(text: str) -> bool:
    matched = sum(1 for keyword in FLOW_KEYWORDS if keyword in (text or ""))
    return matched >= 2


def extract_flow_rows_from_spreadsheet(file_bytes: bytes, suffix: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    if suffix == ".xlsx":
        return _read_xlsx_rows(file_bytes), []
    if suffix == ".xls":
        if xlrd is None:
            return [], ["当前环境未安装 xlrd，暂不支持直接读取 .xls 流水，请优先上传 pdf 或 xlsx。"]
        return _read_xls_rows(file_bytes), []
    return [], [f"暂不支持的流水文件格式: {suffix}"]


def _read_xlsx_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index, header_map = _locate_header_row(rows)
    if header_index is None:
        return []

    parsed_rows = []
    for raw_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row_dict = _build_row_dict_from_tabular_row(row, header_map)
        if not row_dict.get("date") or row_dict.get("amount") in (None, ""):
            continue
        row_dict["row_number"] = raw_index
        parsed_rows.append(row_dict)

    return parsed_rows


def _read_xls_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    rows = [sheet.row_values(i) for i in range(sheet.nrows)]

    header_index, header_map = _locate_header_row(rows)
    if header_index is None:
        return []

    parsed_rows = []
    for raw_index in range(header_index + 1, len(rows)):
        row = rows[raw_index]
        row_dict = _build_row_dict_from_tabular_row(row, header_map, datemode=workbook.datemode)
        if not row_dict.get("date") or row_dict.get("amount") in (None, ""):
            continue
        row_dict["row_number"] = raw_index + 1
        parsed_rows.append(row_dict)

    return parsed_rows


def _locate_header_row(rows: List[Iterable[Any]]) -> Tuple[Optional[int], Dict[str, int]]:
    for idx, row in enumerate(rows[:10]):
        normalized = [_normalize_header_text(item) for item in row]
        header_map: Dict[str, int] = {}
        for col_idx, header in enumerate(normalized):
            if "发生方向" in header or header == "方向":
                header_map["direction"] = col_idx
            elif "交易方向" in header:
                header_map["direction"] = col_idx
            elif "发生类型" in header or header == "类型":
                header_map["subject"] = col_idx
            elif "交易对象" in header:
                header_map["subject"] = col_idx
            elif "实际还款日" in header or header == "日期" or "交易日期" in header:
                header_map["date"] = col_idx
            elif "发生日期" in header:
                header_map["date"] = col_idx
            elif "贷方发生金额" in header:
                header_map["credit_amount"] = col_idx
            elif "借方发生金额" in header:
                header_map["debit_amount"] = col_idx
            elif "发生金额" in header or header == "金额":
                header_map["amount"] = col_idx
            elif "贷款发放日" in header:
                header_map["loan_start_date"] = col_idx
            elif header == "发放日":
                header_map["loan_start_date"] = col_idx
            elif "借据到期日期" in header or "到期日期" in header:
                header_map["contract_end_date"] = col_idx
            elif header == "到期日":
                header_map["contract_end_date"] = col_idx
            elif "还款方式" in header:
                header_map["repayment_mode"] = col_idx
            elif "交易信息" in header or "摘要" in header or "备注" in header:
                header_map["description"] = col_idx
            elif "交易备注" in header:
                header_map["description"] = col_idx
            elif "流水号" in header:
                header_map["serial_no"] = col_idx
            elif "借据号" in header:
                header_map["loan_no"] = col_idx
        if {"direction", "subject", "date"} <= set(header_map) and (
            "amount" in header_map or {"credit_amount", "debit_amount"} <= set(header_map)
        ):
            return idx, header_map
    return None, {}


def _build_row_dict_from_tabular_row(
    row: Iterable[Any],
    header_map: Dict[str, int],
    datemode: int = 0,
) -> Dict[str, Any]:
    values = list(row)

    def value_of(key: str) -> Any:
        idx = header_map.get(key)
        return values[idx] if idx is not None and idx < len(values) else None

    return {
        "serial_no": value_of("serial_no"),
        "loan_no": _stringify(value_of("loan_no")),
        "direction": _normalize_direction(value_of("direction")),
        "subject": _stringify(value_of("subject")),
        "date": _normalize_tabular_date(value_of("date"), datemode),
        "amount": _resolve_tabular_amount(
            direction=_normalize_direction(value_of("direction")),
            amount_value=value_of("amount"),
            credit_value=value_of("credit_amount"),
            debit_value=value_of("debit_amount"),
        ),
        "loan_start_date": _normalize_tabular_date(value_of("loan_start_date"), datemode),
        "contract_end_date": _normalize_tabular_date(value_of("contract_end_date"), datemode),
        "repayment_mode": _stringify(value_of("repayment_mode")),
        "description": _stringify(value_of("description")),
        "raw_line": " | ".join(_stringify(item) for item in values if _stringify(item)),
    }


def extract_flow_rows_from_text(text: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    seen = set()
    for raw_line in (text or "").splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            continue
        if not any(keyword in line for keyword in FLOW_KEYWORDS):
            continue
        date_text = _extract_first_date_text(line)
        amount = _extract_best_amount_from_text_line(line)
        if not date_text or amount is None:
            continue
        description = _extract_description_from_line(line)
        direction = _normalize_direction(line)
        subject = "本金" if "本金" in line else "利息"
        key = (date_text, description, round(amount, 2))
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "serial_no": "",
                "direction": direction,
                "subject": subject,
                "date": date_text,
                "amount": amount,
                "loan_start_date": "",
                "contract_end_date": "",
                "repayment_mode": "",
                "description": description,
                "raw_line": line,
            }
        )
    return rows


def classify_flow_events(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    events = []
    for idx, row in enumerate(rows):
        event_type = _classify_event_type(row)
        if not event_type:
            continue
        events.append(
            {
                "id": idx,
                "date": row["date"],
                "parsed_date": parse_date(row["date"]),
                "amount": round(float(row["amount"]), 2),
                "direction": row.get("direction", ""),
                "subject": row.get("subject", ""),
                "description": row.get("description", ""),
                "raw_line": row.get("raw_line", ""),
                "event_type": event_type,
            }
        )
    return sorted(events, key=lambda item: (item["parsed_date"], item["id"]))


def calculate_loan_flow_result(
    loan_data: Dict[str, Any],
    repayment_records: Optional[List[Dict[str, Any]]] = None,
    rate_adjustments: Optional[List[Dict[str, Any]]] = None,
    calculation_date: Optional[date] = None,
) -> Dict[str, Any]:
    payload = (loan_data or {}).get("_template_payload") or {}
    flow_events = payload.get("events") or []
    if not flow_events:
        raise ValueError("未找到贷款流水事件，无法生成贷款流水计算表")

    loan_amount = Decimal(str(loan_data["loan_amount"]))
    start_date = parse_date(loan_data["start_date"])
    end_date = parse_date(loan_data["end_date"])
    acceleration_date = parse_date(loan_data.get("acceleration_date") or loan_data["end_date"])
    calculation_date = calculation_date or parse_date(loan_data.get("calculation_date") or loan_data["end_date"])

    opening_rate = _normalize_rate(loan_data["annual_interest_rate"])
    expected_remaining_principal = Decimal(str(loan_data.get("remaining_principal", 0) or 0))
    expected_interest_total = Decimal(str(loan_data.get("accrued_interest", 0) or 0))
    expected_penalty_total = Decimal(str(loan_data.get("penalty_interest", 0) or 0))
    expected_compound_total = Decimal(str(loan_data.get("compound_interest", 0) or 0))

    due_principal_map: Dict[date, Decimal] = {}
    for event in flow_events:
        if event["event_type"] != "principal_due_transfer":
            continue
        due_principal_map[event["parsed_date"]] = due_principal_map.get(event["parsed_date"], Decimal("0")) + Decimal(str(event["amount"]))

    for due_date, due_amount in _prepare_manual_principal_plan(loan_data).items():
        due_principal_map[due_date] = due_principal_map.get(due_date, Decimal("0")) + due_amount

    repayments = _prepare_repayment_items(repayment_records or [])
    repayments_by_date: Dict[date, List[Dict[str, Any]]] = {}
    for item in repayments:
        repayments_by_date.setdefault(item["parsed_date"], []).append(item)

    adjustments = _prepare_rate_adjustments(opening_rate, rate_adjustments or [], start_date)
    adjustment_dates = [item["date"] for item in adjustments[1:]]
    interest_cycle_settings = _prepare_interest_cycle_settings(loan_data, start_date)
    interest_cycle_change_dates = [item["effective_date"] for item in interest_cycle_settings[1:]]
    acceleration_start = acceleration_date + timedelta(days=1)
    periods = _build_periods(
        start_date=start_date,
        calculation_date=calculation_date,
        due_principal_map=due_principal_map,
        adjustment_dates=adjustment_dates,
        interest_cycle_change_dates=interest_cycle_change_dates,
        interest_cycle_settings=interest_cycle_settings,
        acceleration_start=acceleration_start,
    )

    current_normal_principal = loan_amount
    current_overdue_principal = Decimal("0")
    current_interest_balance = Decimal("0")
    current_penalty_balance = Decimal("0")
    current_compound_balance = Decimal("0")
    accelerated = False
    consumed_repayment_ids = set()

    detail_rows: List[Dict[str, Any]] = []
    repayment_details: List[Dict[str, Any]] = []

    for seq_no, period in enumerate(periods, start=1):
        period_start = period["start_date"]
        period_end = period["end_date"]
        days = Decimal((period_end - period_start).days + 1)
        normal_rate = _rate_for_date(adjustments, period_start)
        penalty_rate = normal_rate * Decimal("1.5")
        compound_rate = penalty_rate if period_start >= acceleration_start else normal_rate

        opening_normal_principal = current_normal_principal
        opening_overdue_principal = current_overdue_principal
        opening_interest_balance = current_interest_balance
        opening_penalty_balance = current_penalty_balance
        opening_compound_balance = current_compound_balance

        due_principal = Decimal("0")
        structural_principal_repaid = Decimal("0")
        transfer_to_overdue = Decimal("0")
        accelerated_transfer = Decimal("0")

        same_day_repayments = [
            item for item in repayments_by_date.get(period_end, [])
            if item["id"] not in consumed_repayment_ids
        ]

        if period["is_due_row"]:
            due_principal = due_principal_map.get(period_start, Decimal("0"))
            transfer_to_overdue = due_principal
            current_normal_principal = max(current_normal_principal - due_principal, Decimal("0"))
            current_overdue_principal += due_principal

            for repayment in same_day_repayments:
                if repayment["type"] != RepaymentType.PRINCIPAL.value:
                    continue
                applied = min(Decimal(str(repayment["amount"])), current_overdue_principal)
                if applied <= 0:
                    continue
                structural_principal_repaid += applied
                current_overdue_principal -= applied
                consumed_repayment_ids.add(repayment["id"])
                repayment_details.append(
                    _build_repayment_detail(repayment, seq_no, period_start, period_end, principal=applied)
                )

        if period_start >= acceleration_start and not accelerated:
            accelerated_transfer = current_normal_principal
            current_overdue_principal += current_normal_principal
            current_normal_principal = Decimal("0")
            accelerated = True

        normal_interest_base = opening_normal_principal if period_start < acceleration_start else Decimal("0")
        overdue_interest_base = current_overdue_principal

        accrued_interest = _calculate_amount(normal_interest_base, normal_rate, days)
        accrued_penalty = _calculate_amount(overdue_interest_base, penalty_rate, days)
        accrued_compound = _calculate_amount(opening_interest_balance, compound_rate, days)

        current_interest_balance += accrued_interest
        current_penalty_balance += accrued_penalty
        current_compound_balance += accrued_compound

        paid_interest = Decimal("0")
        paid_penalty = Decimal("0")
        paid_compound = Decimal("0")
        extra_principal_paid = Decimal("0")
        repayment_notes: List[str] = []

        next_day = period_end + timedelta(days=1)
        attributed_repayments = list(same_day_repayments)
        if next_day <= calculation_date:
            attributed_repayments.extend(
                item
                for item in repayments_by_date.get(next_day, [])
                if item["id"] not in consumed_repayment_ids and item["type"] in {
                    RepaymentType.INTEREST.value,
                    RepaymentType.PENALTY.value,
                    RepaymentType.COMPOUND.value,
                }
            )

        for repayment in attributed_repayments:
            if repayment["id"] in consumed_repayment_ids:
                continue
            amount = Decimal(str(repayment["amount"]))
            allocated_principal = Decimal("0")
            allocated_interest = Decimal("0")
            allocated_penalty = Decimal("0")
            allocated_compound = Decimal("0")

            if repayment["type"] == RepaymentType.PRINCIPAL.value:
                allocated_principal = min(amount, current_overdue_principal)
                current_overdue_principal -= allocated_principal
                amount -= allocated_principal
                if amount > 0:
                    extra = min(amount, current_normal_principal)
                    current_normal_principal -= extra
                    allocated_principal += extra
                    amount -= extra
                extra_principal_paid += allocated_principal
            elif repayment["type"] == RepaymentType.INTEREST.value:
                allocated_interest = min(amount, current_interest_balance)
                current_interest_balance -= allocated_interest
                paid_interest += allocated_interest
                amount -= allocated_interest
            elif repayment["type"] == RepaymentType.PENALTY.value:
                allocated_penalty = min(amount, current_penalty_balance)
                current_penalty_balance -= allocated_penalty
                paid_penalty += allocated_penalty
                amount -= allocated_penalty
            elif repayment["type"] == RepaymentType.COMPOUND.value:
                allocated_compound = min(amount, current_compound_balance)
                current_compound_balance -= allocated_compound
                paid_compound += allocated_compound
                amount -= allocated_compound

            consumed_repayment_ids.add(repayment["id"])
            repayment_details.append(
                _build_repayment_detail(
                    repayment,
                    seq_no,
                    period_start,
                    period_end,
                    principal=allocated_principal,
                    interest=allocated_interest,
                    penalty=allocated_penalty,
                    compound=allocated_compound,
                    attributed_next_day=(repayment["parsed_date"] == next_day and repayment["type"] != RepaymentType.PRINCIPAL.value),
                )
            )

            repayment_notes.append(
                f"{repayment['date']} {RepaymentType.get_display_name(repayment['type'])} "
                f"{_fmt_money(Decimal(str(repayment['amount'])))}"
            )

        paid_principal_total = structural_principal_repaid + extra_principal_paid
        total_principal_balance = current_normal_principal + current_overdue_principal

        row = {
            "seq_no": seq_no,
            "normal_principal_balance": _to_float(opening_normal_principal),
            "start_date": format_date(period_start),
            "end_date": format_date(period_end),
            "days": int(days),
            "due_principal": _to_float(due_principal),
            "paid_principal": _to_float(paid_principal_total),
            "overdue_principal": _to_float(current_overdue_principal),
            "normal_rate": _to_percent(normal_rate),
            "accrued_interest": _to_float(accrued_interest),
            "paid_interest": _to_float(paid_interest),
            "interest_balance": _to_float(current_interest_balance),
            "penalty_rate": _to_percent(penalty_rate),
            "accrued_penalty": _to_float(accrued_penalty),
            "paid_penalty": _to_float(paid_penalty),
            "penalty_balance": _to_float(current_penalty_balance),
            "compound_rate": _to_percent(compound_rate),
            "accrued_compound": _to_float(accrued_compound),
            "paid_compound": _to_float(paid_compound),
            "compound_balance": _to_float(current_compound_balance),
            "principal_total_balance": _to_float(total_principal_balance),
            "interest_formula": "0.00" if accrued_interest <= 0 else (
                f"{_fmt_money(normal_interest_base)} × {_fmt_percent(normal_rate)} × {int(days)} ÷ 360 = {_fmt_money(accrued_interest)}"
            ),
            "penalty_formula": "0.00" if accrued_penalty <= 0 else (
                f"{_fmt_money(overdue_interest_base)} × {_fmt_percent(penalty_rate)} × {int(days)} ÷ 360 = {_fmt_money(accrued_penalty)}"
            ),
            "compound_formula": "0.00" if accrued_compound <= 0 else (
                f"{_fmt_money(opening_interest_balance)} × {_fmt_percent(compound_rate)} × {int(days)} ÷ 360 = {_fmt_money(accrued_compound)}"
            ),
            "interest_balance_formula": (
                f"{_fmt_money(opening_interest_balance)} + {_fmt_money(accrued_interest)} - {_fmt_money(paid_interest)} = {_fmt_money(current_interest_balance)}"
            ),
            "penalty_balance_formula": (
                f"{_fmt_money(opening_penalty_balance)} + {_fmt_money(accrued_penalty)} - {_fmt_money(paid_penalty)} = {_fmt_money(current_penalty_balance)}"
            ),
            "compound_balance_formula": (
                f"{_fmt_money(opening_compound_balance)} + {_fmt_money(accrued_compound)} - {_fmt_money(paid_compound)} = {_fmt_money(current_compound_balance)}"
            ),
            "remark": _build_row_remark(
                due_principal=due_principal,
                transfer_to_overdue=transfer_to_overdue,
                accelerated_transfer=accelerated_transfer,
                repayment_notes=repayment_notes,
            ),
        }
        detail_rows.append(row)

    summary = {
        "loan_amount": _to_float(loan_amount),
        "remaining_principal": _to_float(current_normal_principal + current_overdue_principal),
        "input_remaining_principal": _to_float(expected_remaining_principal),
        "remaining_principal_difference": _to_float((current_normal_principal + current_overdue_principal) - expected_remaining_principal),
        "accrued_interest": _to_float(current_interest_balance),
        "input_accrued_interest": _to_float(expected_interest_total),
        "accrued_interest_difference": _to_float(current_interest_balance - expected_interest_total),
        "penalty_interest": _to_float(current_penalty_balance),
        "input_penalty_interest": _to_float(expected_penalty_total),
        "penalty_interest_difference": _to_float(current_penalty_balance - expected_penalty_total),
        "compound_interest": _to_float(current_compound_balance),
        "input_compound_interest": _to_float(expected_compound_total),
        "compound_difference": _to_float(current_compound_balance - expected_compound_total),
        "total_amount": _to_float(current_normal_principal + current_overdue_principal + current_interest_balance + current_penalty_balance + current_compound_balance),
        "calculation_date": format_date(calculation_date),
        "template_key": TEMPLATE_KEY,
    }

    return {
        "template_key": TEMPLATE_KEY,
        "template_name": TEMPLATE_NAME,
        "summary": summary,
        "detail": detail_rows,
        "repayment_details": repayment_details,
    }


def _prepare_repayment_items(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    items = []
    for idx, record in enumerate(records):
        items.append(
            {
                "id": idx,
                "date": record["date"],
                "parsed_date": parse_date(record["date"]),
                "type": record.get("type", RepaymentType.INTEREST.value),
                "amount": float(record.get("amount", 0) or 0),
            }
        )
    return sorted(items, key=lambda item: (item["parsed_date"], item["id"]))


def _prepare_rate_adjustments(initial_rate: float, adjustments: List[Dict[str, Any]], start_date: date) -> List[Dict[str, Any]]:
    items = [{"date": start_date, "rate": Decimal(str(initial_rate))}]
    for record in adjustments:
        if not record.get("date") or record.get("rate") in ("", None):
            continue
        items.append(
            {
                "date": parse_date(record["date"]),
                "rate": Decimal(str(_normalize_rate(record["rate"]))),
            }
        )
    items.sort(key=lambda item: item["date"])
    deduped: List[Dict[str, Any]] = []
    for item in items:
        if deduped and deduped[-1]["date"] == item["date"]:
            deduped[-1] = item
        else:
            deduped.append(item)
    return deduped


def _build_periods(
    start_date: date,
    calculation_date: date,
    due_principal_map: Dict[date, Decimal],
    adjustment_dates: List[date],
    interest_cycle_change_dates: List[date],
    interest_cycle_settings: List[Dict[str, Any]],
    acceleration_start: date,
) -> List[Dict[str, Any]]:
    periods: List[Dict[str, Any]] = []
    current = start_date
    due_dates = sorted(item for item in due_principal_map.keys() if start_date <= item <= calculation_date)

    while current <= calculation_date:
        if current in due_principal_map:
            periods.append({"start_date": current, "end_date": current, "is_due_row": True})
            current = current + timedelta(days=1)
            continue

        cycle_months = _interest_cycle_months_for_date(interest_cycle_settings, current)
        nominal_end = _calculate_period_end(current, cycle_months)
        period_end = min(nominal_end, calculation_date)

        next_due = next((item for item in due_dates if current < item <= period_end), None)
        if next_due:
            period_end = min(period_end, next_due - timedelta(days=1))

        next_adjustment = next((item for item in sorted(adjustment_dates) if current < item <= period_end), None)
        if next_adjustment:
            period_end = min(period_end, next_adjustment - timedelta(days=1))

        next_cycle_change = next((item for item in sorted(interest_cycle_change_dates) if current < item <= period_end), None)
        if next_cycle_change:
            period_end = min(period_end, next_cycle_change - timedelta(days=1))

        if current < acceleration_start <= period_end:
            period_end = min(period_end, acceleration_start - timedelta(days=1))

        periods.append({"start_date": current, "end_date": period_end, "is_due_row": False})
        current = period_end + timedelta(days=1)

    return periods


def _rate_for_date(adjustments: List[Dict[str, Any]], target_date: date) -> Decimal:
    current = adjustments[0]["rate"]
    for item in adjustments:
        if item["date"] <= target_date:
            current = item["rate"]
        else:
            break
    return current


def _calculate_amount(base: Decimal, rate: Decimal, days: Decimal) -> Decimal:
    if base <= 0 or rate <= 0 or days <= 0:
        return Decimal("0")
    return base * rate * days / Decimal("360")


def _prepare_manual_principal_plan(loan_data: Dict[str, Any]) -> Dict[date, Decimal]:
    plan_map: Dict[date, Decimal] = {}
    for record in (loan_data or {}).get("principal_plan_records", []) or []:
        if not record or not record.get("date") or record.get("amount") in ("", None):
            continue
        plan_date = parse_date(record["date"])
        plan_amount = Decimal(str(record["amount"]))
        if plan_amount <= 0:
            continue
        plan_map[plan_date] = plan_map.get(plan_date, Decimal("0")) + plan_amount
    return plan_map


def _prepare_interest_cycle_settings(loan_data: Dict[str, Any], start_date: date) -> List[Dict[str, Any]]:
    raw_months = (loan_data or {}).get("interest_cycle_months", 1) or 1
    try:
        initial_months = max(int(raw_months), 1)
    except (TypeError, ValueError):
        initial_months = 1

    items = [{"effective_date": start_date, "months": initial_months}]
    for record in (loan_data or {}).get("interest_cycle_changes", []) or []:
        if not record or not record.get("effective_date") or record.get("months") in ("", None):
            continue
        try:
            months = max(int(record["months"]), 1)
        except (TypeError, ValueError):
            continue
        items.append(
            {
                "effective_date": parse_date(record["effective_date"]),
                "months": months,
            }
        )

    items.sort(key=lambda item: item["effective_date"])
    deduped: List[Dict[str, Any]] = []
    for item in items:
        if deduped and deduped[-1]["effective_date"] == item["effective_date"]:
            deduped[-1] = item
        else:
            deduped.append(item)
    return deduped


def _interest_cycle_months_for_date(settings: List[Dict[str, Any]], target_date: date) -> int:
    current = settings[0]["months"]
    for item in settings:
        if item["effective_date"] <= target_date:
            current = item["months"]
        else:
            break
    return current


def _calculate_period_end(current: date, cycle_months: int) -> date:
    cycle_months = max(int(cycle_months), 1)
    current_month_start = current.replace(day=1)

    if current.day <= 20:
        offset = cycle_months - 1
    else:
        offset = 1 if cycle_months == 1 else cycle_months - 1

    target_month_start = _add_months(current_month_start, offset)
    return date(target_month_start.year, target_month_start.month, 20)


def _add_months(source: date, months: int) -> date:
    month_index = source.month - 1 + months
    year = source.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def _build_row_remark(
    due_principal: Decimal,
    transfer_to_overdue: Decimal,
    accelerated_transfer: Decimal,
    repayment_notes: List[str],
) -> str:
    parts = []
    if due_principal > 0:
        parts.append(f"本期应还本金 {_fmt_money(due_principal)}")
    if transfer_to_overdue > 0:
        parts.append(f"到期转逾期 {_fmt_money(transfer_to_overdue)}")
    if accelerated_transfer > 0:
        parts.append(f"提前到期转逾期 {_fmt_money(accelerated_transfer)}")
    if repayment_notes:
        parts.append("；".join(repayment_notes))
    return "；".join(parts)


def _build_repayment_detail(
    repayment: Dict[str, Any],
    seq_no: int,
    period_start: date,
    period_end: date,
    principal: Decimal = Decimal("0"),
    interest: Decimal = Decimal("0"),
    penalty: Decimal = Decimal("0"),
    compound: Decimal = Decimal("0"),
    attributed_next_day: bool = False,
) -> Dict[str, Any]:
    source = "次日归属上期" if attributed_next_day else "当期内还款"
    return {
        "date": repayment["date"],
        "type": repayment["type"],
        "amount": repayment["amount"],
        "attributed_period_seq": seq_no,
        "attributed_period": f"{format_date(period_start)} 至 {format_date(period_end)}",
        "source": source,
        "allocation": {
            "principal": _to_float(principal),
            "interest": _to_float(interest),
            "penalty": _to_float(penalty),
            "compound_interest": _to_float(compound),
            "compound_penalty": 0.0,
            "unapplied": 0.0,
        },
    }


def _classify_event_type(row: Dict[str, Any]) -> Optional[str]:
    description = (row.get("description") or "").strip()
    direction = row.get("direction", "")
    subject = row.get("subject", "")

    if "放款" in description:
        return "disbursement"
    if "未到期本金到期转桶子" in description:
        return "principal_due_transfer"
    if "罚息结息" in description:
        return "penalty_accrual"
    if "正常本金结息" in description:
        return "normal_interest_accrual"
    if "现金回收" in description and "本金" in description:
        return "principal_repayment"
    if "现金回收" in description and "罚息" in description:
        return "penalty_repayment"
    if "现金回收" in description and "复利" in description:
        return "compound_repayment"
    if "现金回收" in description and ("利息" in description or subject == "利息"):
        return "interest_repayment"

    if direction == "借方" and "本金" in subject:
        return "disbursement"
    if direction == "贷方" and "本金" in subject:
        return "principal_repayment"
    if direction == "贷方" and "利息" in subject:
        return "interest_repayment"
    return None


def _event_type_to_repayment_type(event_type: str) -> Optional[str]:
    mapping = {
        "principal_repayment": RepaymentType.PRINCIPAL.value,
        "interest_repayment": RepaymentType.INTEREST.value,
        "penalty_repayment": RepaymentType.PENALTY.value,
        "compound_repayment": RepaymentType.COMPOUND.value,
    }
    return mapping.get(event_type)


def _normalize_header_text(value: Any) -> str:
    return re.sub(r"\s+", "", _stringify(value))


def _normalize_tabular_date(value: Any, datemode: int = 0) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return format_date(value.date())
    if isinstance(value, date):
        return format_date(value)
    if isinstance(value, (int, float)) and xlrd is not None and value > 1000:
        try:
            return format_date(xlrd.xldate_as_datetime(value, datemode).date())
        except Exception:
            return ""
    text = _stringify(value)
    return _normalize_date_text(text)


def _normalize_date_text(text: str) -> str:
    date_text = _extract_first_date_text(text)
    return date_text or ""


def _extract_first_date_text(text: str) -> Optional[str]:
    if not text:
        return None
    match = re.search(r"(20\d{2})[年\-/\.](\d{1,2})[月\-/\.](\d{1,2})日?", str(text))
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def _normalize_direction(value: Any) -> str:
    text = _stringify(value)
    if "借" in text:
        return "借方"
    if "贷" in text:
        return "贷方"
    return text


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_amount(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return abs(float(value))
    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return None
    try:
        return abs(float(text))
    except ValueError:
        return None


def _resolve_tabular_amount(
    direction: str,
    amount_value: Any,
    credit_value: Any,
    debit_value: Any,
) -> Optional[float]:
    direct_amount = _normalize_amount(amount_value)
    if direct_amount is not None:
        return direct_amount

    credit_amount = _normalize_amount(credit_value)
    debit_amount = _normalize_amount(debit_value)

    if direction == "借方" and debit_amount not in (None, 0):
        return debit_amount
    if direction == "贷方" and credit_amount not in (None, 0):
        return credit_amount
    if debit_amount not in (None, 0):
        return debit_amount
    if credit_amount not in (None, 0):
        return credit_amount
    return None


def _extract_best_amount_from_text_line(line: str) -> Optional[float]:
    without_dates = re.sub(r"20\d{2}[年\-/\.]\d{1,2}[月\-/\.]\d{1,2}日?", " ", line)
    matches = list(re.finditer(r"\d[\d,]*\.\d+|\d[\d,]*", without_dates))
    amounts: List[Tuple[int, float]] = []
    for match in matches:
        value = _normalize_amount(match.group())
        if value is None:
            continue
        if 1900 <= value <= 2100:
            continue
        amounts.append((match.start(), value))
    if not amounts:
        return None
    for _, amount in amounts:
        if amount >= 1:
            return amount
    return amounts[0][1]


def _extract_description_from_line(line: str) -> str:
    for keyword in FLOW_KEYWORDS:
        if keyword in line:
            return keyword
    if "现金回收" in line:
        return "现金回收"
    return line[:40]


def _first_event_amount(events: List[Dict[str, Any]], event_type: str) -> Optional[float]:
    item = next((item for item in events if item["event_type"] == event_type), None)
    return None if item is None else float(item["amount"])


def _first_event_date(events: List[Dict[str, Any]], event_type: str) -> Optional[str]:
    item = next((item for item in events if item["event_type"] == event_type), None)
    return None if item is None else item["date"]


def _last_event_date(events: List[Dict[str, Any]]) -> Optional[str]:
    if not events:
        return None
    return events[-1]["date"]


def _first_non_empty(rows: List[Dict[str, Any]], key: str) -> str:
    for row in rows:
        value = row.get(key)
        if value not in ("", None):
            return str(value)
    return ""


def _select_relevant_loan_rows(rows: List[Dict[str, Any]], filename: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    if not rows:
        return rows, []

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        key = (
            row.get("loan_no")
            or f"{row.get('loan_start_date', '')}|{row.get('contract_end_date', '')}"
            or "__default__"
        )
        grouped.setdefault(str(key), []).append(row)

    if len(grouped) <= 1:
        return rows, []

    amount_hint = _extract_amount_hint_from_filename(filename)
    best_key = None
    best_score = None

    for key, group_rows in grouped.items():
        events = classify_flow_events(group_rows)
        disbursement = next((item for item in events if item["event_type"] == "disbursement"), None)
        disbursement_amount = float(disbursement["amount"]) if disbursement else 0.0
        disbursement_date = disbursement["parsed_date"].toordinal() if disbursement else 0

        score = (
            1 if amount_hint and abs(disbursement_amount - amount_hint) < 0.01 else 0,
            1 if disbursement else 0,
            disbursement_date,
            len(group_rows),
        )
        if best_score is None or score > best_score:
            best_key = key
            best_score = score

    selected_rows = grouped[best_key] if best_key is not None else rows
    warnings = []
    if best_key is not None:
        selected_disbursement = next(
            (item for item in classify_flow_events(selected_rows) if item["event_type"] == "disbursement"),
            None,
        )
        description = f"借据号 {best_key}"
        if selected_disbursement:
            description += f"，放款金额 {selected_disbursement['amount']}"
        if amount_hint:
            warnings.append(f"检测到文件内存在多笔借据，已按文件名金额优先选择 {description}。")
        else:
            warnings.append(f"检测到文件内存在多笔借据，已自动选择 {description}。")
    return selected_rows, warnings


def _extract_amount_hint_from_filename(filename: str) -> Optional[float]:
    match = re.search(r"(\d+(?:\.\d+)?)\s*万", filename or "")
    if not match:
        return None
    return float(match.group(1)) * 10000


def _normalize_rate(value: Any) -> float:
    if isinstance(value, str):
        value = value.replace("%", "").strip()
    rate = float(value)
    return rate / 100 if rate > 1 else rate


def _to_percent(value: Decimal) -> float:
    return round(float(value * Decimal("100")), 4)


def _fmt_percent(value: Decimal) -> str:
    return f"{_to_percent(value):.4f}%"


def _fmt_money(value: Decimal) -> str:
    quantized = Decimal(str(value)).quantize(Decimal("1.000000"), rounding=ROUND_HALF_UP)
    text = f"{quantized:.6f}"
    return text.rstrip("0").rstrip(".") if "." in text else text


def _to_float(value: Decimal) -> float:
    return float(Decimal(str(value)).quantize(Decimal("1.000000"), rounding=ROUND_HALF_UP))
